#!/usr/bin/env python3
"""
NSN Intelligence Suite - Flask Web Application
===============================================
Single-file backend replacing the Excel-based pipeline system.
All pipeline logic ported from the original Python scripts.
Data stored in SQLite (nsn_intel.db).
"""
from __future__ import annotations
import sqlite3, os, math, json, re, csv, statistics, logging, shutil, threading, time, io
from pathlib import Path
from datetime import datetime, timezone
from collections import defaultdict
from typing import Dict, List, Tuple, Optional

from flask import Flask, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename
import requests as _requests_lib
from bs4 import BeautifulSoup as _BS4

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500 MB

BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / "nsn_intel.db"
UPLOAD_DIR = BASE_DIR / "uploads"
PDF_DIR = UPLOAD_DIR / "pdfs"
CRAWLER_DL_DIR = BASE_DIR / "crawler_downloads"

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s %(message)s")
log = logging.getLogger("nsn_app")

# ──────────────────────────────────────────────
# SECTION A: Database
# ──────────────────────────────────────────────

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS niins (
    niin TEXT PRIMARY KEY,
    nsn TEXT DEFAULT '',
    notes TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS niin_targets (
    niin TEXT PRIMARY KEY,
    added_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS command_center (
    niin TEXT PRIMARY KEY,
    coverage REAL DEFAULT 0,
    docs_total INTEGER DEFAULT 0,
    fp_conf REAL DEFAULT 0,
    nobid_score REAL DEFAULT 0,
    opportunity_score REAL DEFAULT 0,
    action TEXT DEFAULT '',
    confidence REAL DEFAULT 0,
    primary_reason TEXT DEFAULT '',
    next_steps TEXT DEFAULT '',
    evidence_score REAL DEFAULT 0,
    fp_score REAL DEFAULT 0,
    docs_score REAL DEFAULT 0,
    profit_score_val REAL DEFAULT 0,
    best_url TEXT DEFAULT '',
    top_gaps TEXT DEFAULT '',
    pn_dup_count INTEGER DEFAULT 0,
    median_unit_price REAL,
    supplier_count INTEGER DEFAULT 0,
    updated_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS pricing_signals (
    niin TEXT PRIMARY KEY,
    last_unit_price REAL,
    median_unit_price REAL,
    min_unit_price REAL,
    max_unit_price REAL,
    total_awards INTEGER DEFAULT 0,
    total_qty INTEGER DEFAULT 0,
    last_award_date TEXT,
    supplier_count INTEGER DEFAULT 0,
    top_supplier TEXT DEFAULT '',
    source_file TEXT DEFAULT '',
    notes TEXT DEFAULT '',
    imported_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS nobid_signals (
    niin TEXT PRIMARY KEY,
    nobid_count INTEGER DEFAULT 0,
    last_close TEXT DEFAULT '',
    first_close TEXT DEFAULT '',
    total_pr_value REAL DEFAULT 0,
    fsc TEXT DEFAULT '',
    nomenclature TEXT DEFAULT '',
    imported_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS doc_extract (
    doc_id TEXT PRIMARY KEY,
    filename TEXT NOT NULL,
    doc_type TEXT DEFAULT 'PDF',
    niin TEXT DEFAULT '',
    nsn TEXT DEFAULT '',
    pn_candidates TEXT DEFAULT '',
    key_facts TEXT DEFAULT '',
    extract_json TEXT DEFAULT '{}',
    confidence REAL DEFAULT 0,
    notes TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS fingerprint_by_niin (
    niin TEXT PRIMARY KEY,
    nsn TEXT DEFAULT '',
    fingerprint_json TEXT DEFAULT '{}',
    evidence_docs TEXT DEFAULT '',
    confidence REAL DEFAULT 0,
    top_facts TEXT DEFAULT '',
    updated_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS flis_refnums (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    niin TEXT DEFAULT '',
    nsn TEXT DEFAULT '',
    cage TEXT DEFAULT '',
    oem_pn TEXT DEFAULT '',
    oem_name TEXT DEFAULT '',
    source_file TEXT DEFAULT '',
    source_row INTEGER DEFAULT 0,
    imported_at TEXT DEFAULT (datetime('now')),
    notes TEXT DEFAULT ''
);

CREATE TABLE IF NOT EXISTS flis_detected_schema (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    source_file TEXT,
    sheet_name TEXT DEFAULT '',
    col_index INTEGER,
    col_letter TEXT,
    header_name TEXT,
    best_guess TEXT,
    score REAL,
    explanation TEXT,
    sample_values TEXT,
    recommended_key TEXT,
    detected_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS evidence_links (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    niin TEXT DEFAULT '',
    url TEXT NOT NULL,
    title TEXT DEFAULT '',
    source TEXT DEFAULT '',
    added_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS config (
    key TEXT NOT NULL,
    value TEXT DEFAULT '',
    section TEXT DEFAULT 'general',
    description TEXT DEFAULT '',
    updated_at TEXT DEFAULT (datetime('now')),
    PRIMARY KEY (key, section)
);

CREATE TABLE IF NOT EXISTS pipeline_log (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    pipeline_name TEXT NOT NULL,
    started_at TEXT DEFAULT (datetime('now')),
    finished_at TEXT,
    rows_processed INTEGER DEFAULT 0,
    status TEXT DEFAULT 'running',
    log_text TEXT DEFAULT ''
);

CREATE TABLE IF NOT EXISTS forecast (
    niin TEXT PRIMARY KEY,
    fsc TEXT DEFAULT '',
    supply_chain TEXT DEFAULT '',
    item_description TEXT DEFAULT '',
    ui TEXT DEFAULT '',
    total_demand INTEGER DEFAULT 0,
    monthly_demand TEXT DEFAULT '{}',
    avg_monthly REAL DEFAULT 0,
    peak_monthly INTEGER DEFAULT 0,
    demand_months INTEGER DEFAULT 0,
    forecast_start TEXT DEFAULT '',
    forecast_end TEXT DEFAULT '',
    source_file TEXT DEFAULT '',
    imported_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS publog_identification (
    niin TEXT PRIMARY KEY,
    fsc TEXT DEFAULT '',
    inc TEXT DEFAULT '',
    item_name TEXT DEFAULT '',
    sos TEXT DEFAULT '',
    end_item_name TEXT DEFAULT '',
    cancelled_niin TEXT DEFAULT '',
    fiig TEXT DEFAULT '',
    crit_cd TEXT DEFAULT '',
    dmil TEXT DEFAULT '',
    pmic TEXT DEFAULT '',
    hmic TEXT DEFAULT '',
    niin_asgmt TEXT DEFAULT '',
    imported_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS publog_characteristics (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    niin TEXT NOT NULL,
    mrc TEXT DEFAULT '',
    requirements_statement TEXT DEFAULT '',
    clear_text_reply TEXT DEFAULT '',
    imported_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS publog_management (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    niin TEXT NOT NULL,
    effective_date TEXT DEFAULT '',
    moe TEXT DEFAULT '',
    aac TEXT DEFAULT '',
    sos TEXT DEFAULT '',
    ui TEXT DEFAULT '',
    ui_conv_fac TEXT DEFAULT '',
    unit_price TEXT DEFAULT '',
    qup TEXT DEFAULT '',
    ciic TEXT DEFAULT '',
    slc TEXT DEFAULT '',
    usc TEXT DEFAULT '',
    imported_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS publog_references (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    niin TEXT NOT NULL,
    part_number TEXT DEFAULT '',
    cage_code TEXT DEFAULT '',
    cage_status TEXT DEFAULT '',
    rncc TEXT DEFAULT '',
    rnvc TEXT DEFAULT '',
    dac TEXT DEFAULT '',
    rnaac TEXT DEFAULT '',
    imported_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS publog_cage (
    cage_code TEXT PRIMARY KEY,
    cage_status TEXT DEFAULT '',
    type TEXT DEFAULT '',
    company TEXT DEFAULT '',
    city TEXT DEFAULT '',
    state_province TEXT DEFAULT '',
    zip_postal_zone TEXT DEFAULT '',
    country TEXT DEFAULT '',
    imported_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS publog_hseries_fsc (
    fsc TEXT PRIMARY KEY,
    fsc_title TEXT DEFAULT '',
    fsg_title TEXT DEFAULT '',
    imported_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS publog_hseries_inc (
    inc TEXT PRIMARY KEY,
    item_name TEXT DEFAULT '',
    fiig TEXT DEFAULT '',
    fsc TEXT DEFAULT '',
    definition TEXT DEFAULT '',
    imported_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS crawler_sites (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    rank INTEGER DEFAULT 0,
    site_name TEXT DEFAULT '',
    url TEXT DEFAULT '',
    primary_value TEXT DEFAULT '',
    expected_roi TEXT DEFAULT '',
    enabled INTEGER DEFAULT 1,
    imported_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS crawler_results (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    niin TEXT NOT NULL,
    site_name TEXT DEFAULT '',
    source_url TEXT DEFAULT '',
    page_url TEXT DEFAULT '',
    page_title TEXT DEFAULT '',
    snippet TEXT DEFAULT '',
    found_pns TEXT DEFAULT '',
    found_specs TEXT DEFAULT '',
    relevance_score REAL DEFAULT 0,
    crawled_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS crawler_files (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    niin TEXT NOT NULL,
    crawler_result_id INTEGER,
    filename TEXT NOT NULL,
    file_path TEXT NOT NULL,
    file_type TEXT DEFAULT 'unknown',
    file_size INTEGER DEFAULT 0,
    mime_type TEXT DEFAULT '',
    source_url TEXT DEFAULT '',
    downloaded_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (crawler_result_id) REFERENCES crawler_results(id)
);

CREATE TABLE IF NOT EXISTS crawler_file_extracts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    file_id INTEGER NOT NULL,
    extract_type TEXT DEFAULT 'text',
    content TEXT DEFAULT '',
    niin_mentions TEXT DEFAULT '',
    pn_mentions TEXT DEFAULT '',
    extracted_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (file_id) REFERENCES crawler_files(id)
);

CREATE TABLE IF NOT EXISTS cad_jobs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    niin TEXT NOT NULL,
    status TEXT DEFAULT 'pending',
    spec_json TEXT DEFAULT '{}',
    scad_code TEXT DEFAULT '',
    cadquery_code TEXT DEFAULT '',
    output_dir TEXT DEFAULT '',
    error_text TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now')),
    finished_at TEXT
);

CREATE INDEX IF NOT EXISTS idx_cc_action ON command_center(action);
CREATE INDEX IF NOT EXISTS idx_cc_score ON command_center(opportunity_score DESC);
CREATE INDEX IF NOT EXISTS idx_doc_niin ON doc_extract(niin);
CREATE INDEX IF NOT EXISTS idx_flis_niin ON flis_refnums(niin);
CREATE INDEX IF NOT EXISTS idx_ev_niin ON evidence_links(niin);
CREATE INDEX IF NOT EXISTS idx_forecast_demand ON forecast(total_demand DESC);
CREATE INDEX IF NOT EXISTS idx_forecast_fsc ON forecast(fsc);
CREATE INDEX IF NOT EXISTS idx_publog_char_niin ON publog_characteristics(niin);
CREATE INDEX IF NOT EXISTS idx_publog_mgmt_niin ON publog_management(niin);
CREATE INDEX IF NOT EXISTS idx_publog_ref_niin ON publog_references(niin);
CREATE INDEX IF NOT EXISTS idx_publog_ref_cage ON publog_references(cage_code);
CREATE INDEX IF NOT EXISTS idx_crawler_results_niin ON crawler_results(niin);
CREATE INDEX IF NOT EXISTS idx_crawler_files_niin ON crawler_files(niin);
CREATE INDEX IF NOT EXISTS idx_crawler_files_result ON crawler_files(crawler_result_id);
CREATE INDEX IF NOT EXISTS idx_cfe_file ON crawler_file_extracts(file_id);
CREATE INDEX IF NOT EXISTS idx_cad_jobs_niin ON cad_jobs(niin);

CREATE TABLE IF NOT EXISTS profit_assessments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    niin TEXT NOT NULL,
    item_name TEXT DEFAULT '',
    govt_unit_price REAL,
    cots_alternatives TEXT DEFAULT '[]',
    best_cots_price REAL,
    best_cots_source TEXT DEFAULT '',
    potential_margin REAL,
    roi_percent REAL,
    annual_demand INTEGER DEFAULT 0,
    annual_revenue_potential REAL,
    assessment_notes TEXT DEFAULT '',
    generated_at TEXT DEFAULT (datetime('now')),
    UNIQUE(niin)
);
CREATE INDEX IF NOT EXISTS idx_profit_niin ON profit_assessments(niin);
"""


def get_db():
    db = sqlite3.connect(str(DB_PATH), timeout=30)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA foreign_keys=ON")
    return db


def _worker_db():
    """Get a DB connection for background worker threads (with WAL mode)."""
    db = sqlite3.connect(str(DB_PATH), timeout=30)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    return db


def init_db():
    db = get_db()
    db.executescript(SCHEMA_SQL)
    # Migrate: add new decision engine columns to command_center
    existing = {r[1] for r in db.execute("PRAGMA table_info(command_center)").fetchall()}
    migrations = [
        ("demand_score", "REAL DEFAULT 0"), ("char_score", "REAL DEFAULT 0"),
        ("supply_score", "REAL DEFAULT 0"), ("price_score", "REAL DEFAULT 0"),
        ("competition_score", "REAL DEFAULT 0"), ("cots_readiness", "REAL DEFAULT 0"),
        ("total_demand", "INTEGER DEFAULT 0"), ("char_count", "INTEGER DEFAULT 0"),
        ("ref_count", "INTEGER DEFAULT 0"), ("cage_count", "INTEGER DEFAULT 0"),
        ("govt_unit_price", "REAL"), ("item_name", "TEXT DEFAULT ''"),
        ("crit_cd", "TEXT DEFAULT ''"), ("dmil", "TEXT DEFAULT ''"),
    ]
    for col, typedef in migrations:
        if col not in existing:
            db.execute(f"ALTER TABLE command_center ADD COLUMN {col} {typedef}")
    # Migrate cad_jobs: add image_path, dxf_path, job_type columns
    cad_cols = {r[1] for r in db.execute("PRAGMA table_info(cad_jobs)").fetchall()}
    for col, typedef in [("image_path", "TEXT DEFAULT ''"), ("dxf_path", "TEXT DEFAULT ''"),
                         ("job_type", "TEXT DEFAULT 'cad'")]:
        if col not in cad_cols:
            db.execute(f"ALTER TABLE cad_jobs ADD COLUMN {col} {typedef}")
    db.commit()
    db.close()


# ──────────────────────────────────────────────
# SECTION B: Utility functions (ported verbatim)
# ──────────────────────────────────────────────

def now_iso():
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def clamp(x, a=0, b=100):
    return max(a, min(b, x))


def to_float(v):
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def to_int(v):
    if v in (None, ""):
        return 0
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return 0


def norm_niin(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    s = re.sub(r"\.0$", "", s)
    s = re.sub(r"\D", "", s)
    return s if len(s) == 9 else None


# ──────────────────────────────────────────────
# SECTION C: Decision Engine (from decision_engine_v3.py)
# ──────────────────────────────────────────────

DEFAULTS = {
    # Opportunity weights (must sum to 1.0)
    "W_DEMAND": 0.30, "W_CHARS": 0.25, "W_SUPPLY": 0.15,
    "W_PRICE": 0.15, "W_COMPETITION": 0.15,
    # COTS readiness weights
    "CW_CHARS": 0.40, "CW_SUPPLY": 0.30, "CW_DMIL": 0.15, "CW_CRIT": 0.15,
    # Action thresholds
    "BID_NOW_DEMAND": 60, "BID_NOW_PRICE": 50,
    "BID_NOW_SUPPLY": 40, "BID_NOW_COTS": 60,
    "COTS_READINESS_MIN": 70, "COTS_DEMAND_MIN": 30,
    "SAR_COMPETITION": 60, "SAR_CHARS": 40, "SAR_DEMAND": 30,
    "RESEARCH_DEMAND": 20, "RESEARCH_CHARS_LOW": 30, "RESEARCH_SUPPLY_LOW": 20,
    "PASS_DEMAND_MAX": 20, "PASS_PRICE_LOW": 15, "PASS_COMP_LOW": 30,
}

DECISION_DESCRIPTIONS = {
    "W_DEMAND": "Opportunity weight for forecast demand",
    "W_CHARS": "Opportunity weight for PUBLOG characteristics count",
    "W_SUPPLY": "Opportunity weight for known suppliers/part numbers",
    "W_PRICE": "Opportunity weight for unit price (profit opportunity)",
    "W_COMPETITION": "Opportunity weight for competition level (no-bid + few suppliers)",
    "CW_CHARS": "COTS readiness weight for characteristics completeness",
    "CW_SUPPLY": "COTS readiness weight for known OEM sources",
    "CW_DMIL": "COTS readiness weight for demil restriction check",
    "CW_CRIT": "COTS readiness weight for criticality code check",
    "BID_NOW_DEMAND": "Min demand score for BID_NOW action",
    "BID_NOW_PRICE": "Min price score for BID_NOW action",
    "BID_NOW_SUPPLY": "Min supply score for BID_NOW action",
    "BID_NOW_COTS": "Min COTS readiness for BID_NOW action",
    "COTS_READINESS_MIN": "Min COTS readiness score for COTS_VALIDATE",
    "COTS_DEMAND_MIN": "Min demand score for COTS_VALIDATE",
    "SAR_COMPETITION": "Min competition score for SAR_TARGET",
    "SAR_CHARS": "Min characteristics score for SAR_TARGET",
    "SAR_DEMAND": "Min demand score for SAR_TARGET",
    "RESEARCH_DEMAND": "Min demand score to qualify for RESEARCH",
    "RESEARCH_CHARS_LOW": "Chars score below this triggers RESEARCH",
    "RESEARCH_SUPPLY_LOW": "Supply score below this triggers RESEARCH",
    "PASS_DEMAND_MAX": "Demand score at or below this => PASS",
    "PASS_PRICE_LOW": "Price score below this (with low competition) => PASS",
    "PASS_COMP_LOW": "Competition score below this (with low price) => PASS",
}

FINGERPRINT_DEFAULTS = {
    "MAX_PAGES_PER_PDF": "12",
    "MAX_CHARS_PER_PDF": "180000",
    "PDF_DIR": "",
    "DOC_TYPE_RULES": "datasheet=OEM_DATASHEET;TM_=MIL_TM;SPE=RFQ",
    "MATERIAL_KEYWORDS": "stainless;steel;aluminum;brass;bronze;titanium;inconel;monel;hastelloy;copper;nickel;rubber;teflon;nylon;ptfe;viton;buna;silicone;ceramic;carbon",
    "VALVE_TYPE_KEYWORDS": "check valve;relief valve;shuttle valve;ball valve;gate valve;globe valve;needle valve;solenoid;regulator;actuator;fitting;coupling;adapter;connector;manifold",
    "THREAD_REGEX": r"\b(AN\s?\d+|MS\s?\d+|\d{1,2}\s?[-/]\s?\d{1,2}\s?(UNF|UNC|UNJF|UNJC|NPT|NPTF|SAE|JIC|ORB|ORFS))\b",
    "PRESSURE_REGEX": r"(\d{1,6}(?:\.\d+)?)\s*(?:psig?|PSI(?:G)?)",
    "TEMP_REGEX_F": r"(-?\d{1,4}(?:\.\d+)?)\s*(?:\u00b0\s*F|deg\s*F|degrees?\s*F)",
    "STD_REGEX": r"\b(MIL-[A-Z]-\d+[A-Z]?|SAE\s*(?:AS|AMS|J)?\d+|NAS\s*\d+|AN\s*\d+|ASTM\s*[A-Z]\d+|AMS\s*\d+)\b",
}

FINGERPRINT_DESCRIPTIONS = {
    "MAX_PAGES_PER_PDF": "Max pages to scan per PDF",
    "MAX_CHARS_PER_PDF": "Max characters to extract per PDF",
    "PDF_DIR": "Override PDF directory (leave blank for uploads/pdfs)",
    "DOC_TYPE_RULES": "Filename pattern to doc-type mapping (semicolon-separated key=value)",
    "MATERIAL_KEYWORDS": "Semicolon-separated material keywords to detect",
    "VALVE_TYPE_KEYWORDS": "Semicolon-separated valve/item type keywords",
    "THREAD_REGEX": "Regex for thread specification extraction",
    "PRESSURE_REGEX": "Regex for pressure value extraction (group 1 = number)",
    "TEMP_REGEX_F": "Regex for Fahrenheit temperature extraction (group 1 = number)",
    "STD_REGEX": "Regex for military/industry standard extraction",
}


CRAWLER_DEFAULTS = {
    "MAX_FILE_SIZE_MB": 50,
    "MAX_FILES_PER_NIIN": 20,
    "MIN_RELEVANCE_FOR_DOWNLOAD": 20,
    "DOWNLOAD_PDFS": 1,
    "DOWNLOAD_IMAGES": 1,
    "DOWNLOAD_DOCS": 1,
}
CRAWLER_DESCRIPTIONS = {
    "MAX_FILE_SIZE_MB": "Maximum file size to download in MB",
    "MAX_FILES_PER_NIIN": "Maximum number of files to download per NIIN",
    "MIN_RELEVANCE_FOR_DOWNLOAD": "Minimum relevance score to trigger file downloads",
    "DOWNLOAD_PDFS": "Download PDF files (1=yes, 0=no)",
    "DOWNLOAD_IMAGES": "Download image files (1=yes, 0=no)",
    "DOWNLOAD_DOCS": "Download document files (1=yes, 0=no)",
}

# Regex patterns for document intelligence extraction
NIIN_PATTERN = re.compile(r'\b\d{9}\b')
NSN_PATTERN = re.compile(r'\b\d{4}-\d{2}-\d{3}-\d{4}\b')
PN_PATTERN = re.compile(r'\b[A-Z0-9]{2,}[-/][A-Z0-9]{2,}(?:[-/][A-Z0-9]+)*\b')


def load_config_from_db(db, section: str, defaults: dict) -> dict:
    cfg = dict(defaults)
    rows = db.execute("SELECT key, value FROM config WHERE section=?", (section,)).fetchall()
    for r in rows:
        k, v = r["key"], r["value"]
        if k in cfg and v not in (None, ""):
            orig = defaults.get(k)
            if isinstance(orig, float):
                try:
                    cfg[k] = float(v)
                except (ValueError, TypeError):
                    pass
            elif isinstance(orig, int):
                try:
                    cfg[k] = int(float(v))
                except (ValueError, TypeError):
                    pass
            else:
                cfg[k] = v
    return cfg


def calc_demand_score(total_demand):
    """0 demand→0, 10→25, 100→50, 1000→75, 10000→100"""
    return clamp(math.log10(1 + max(0, total_demand or 0)) * 25)

def calc_char_score(char_count):
    """0 chars→0, 5→25, 10→50, 20→100"""
    return clamp((char_count or 0) * 5)

def calc_supply_score(unique_cages, part_number_count):
    """More OEM sources = more COTS potential"""
    return clamp((unique_cages or 0) * 15 + (part_number_count or 0) * 5)

def calc_price_score(unit_price):
    """Higher unit price = more profit opportunity"""
    if unit_price is None:
        return 0
    try:
        p = float(unit_price)
    except (ValueError, TypeError):
        return 0
    return clamp(math.log10(1 + max(0, p)) * 25)

def calc_competition_score(award_supplier_count, nobid_count):
    """Low competition = higher opportunity"""
    return clamp(100 - min(100, (award_supplier_count or 0) * 20) + (nobid_count or 0) * 10)

def calc_cots_readiness(char_score, supply_score, dmil, crit_cd, cfg):
    """COTS viability: well-defined + known sources + no restrictions"""
    dmil_ok = 100 if not dmil or dmil.strip() == '' else 0
    crit_ok = 100 if crit_cd not in ('C', 'D') else 0
    return clamp(
        cfg["CW_CHARS"] * char_score +
        cfg["CW_SUPPLY"] * supply_score +
        cfg["CW_DMIL"] * dmil_ok +
        cfg["CW_CRIT"] * crit_ok
    )

def decide(demand_s, char_s, supply_s, price_s, comp_s, cots_r,
           has_forecast, has_chars, has_pricing, has_refs, has_mgmt, cfg):
    """New decision engine using forecast + PUBLOG data."""
    opp = clamp(
        cfg["W_DEMAND"] * demand_s + cfg["W_CHARS"] * char_s +
        cfg["W_SUPPLY"] * supply_s + cfg["W_PRICE"] * price_s +
        cfg["W_COMPETITION"] * comp_s
    )
    # Confidence = data completeness
    conf = clamp(
        (has_forecast * 0.25 + has_chars * 0.25 + has_pricing * 0.20 +
         has_refs * 0.15 + has_mgmt * 0.15), 0, 1
    )
    subs = {"demand": round(demand_s, 1), "chars": round(char_s, 1),
            "supply": round(supply_s, 1), "price": round(price_s, 1),
            "competition": round(comp_s, 1), "cots_readiness": round(cots_r, 1)}

    # Priority cascade
    if (demand_s >= cfg["BID_NOW_DEMAND"] and price_s >= cfg["BID_NOW_PRICE"] and
            supply_s >= cfg["BID_NOW_SUPPLY"] and cots_r >= cfg["BID_NOW_COTS"]):
        return opp, "BID_NOW", round(conf, 2), \
            "High demand + good price + known suppliers + COTS-viable.", \
            "Pull 1-3 quotes; validate FFF against characteristics; prep bid package.", subs

    if cots_r >= cfg["COTS_READINESS_MIN"] and demand_s >= cfg["COTS_DEMAND_MIN"]:
        return opp, "COTS_VALIDATE", round(conf, 2), \
            "Strong COTS potential — well-defined item with known sources.", \
            "Shortlist COTS candidates; compare specs to characteristics; get quotes.", subs

    if (comp_s >= cfg["SAR_COMPETITION"] and char_s >= cfg["SAR_CHARS"] and
            demand_s >= cfg["SAR_DEMAND"]):
        return opp, "SAR_TARGET", round(conf, 2), \
            "Low competition + defined item — SAR opportunity.", \
            "Assemble SAR evidence; map spec deltas; define test/inspection plan.", subs

    if demand_s >= cfg["RESEARCH_DEMAND"] and (char_s < cfg["RESEARCH_CHARS_LOW"] or
                                                supply_s < cfg["RESEARCH_SUPPLY_LOW"]):
        return opp, "RESEARCH", round(conf, 2), \
            "Has demand but poorly defined or no known sources.", \
            "Crawl technical sources; gather datasheets; identify OEM part numbers.", subs

    if demand_s < cfg["PASS_DEMAND_MAX"]:
        return opp, "PASS", round(conf, 2), \
            "Low demand — not worth pursuing at this time.", \
            "Monitor for demand changes; revisit if forecast increases.", subs

    if price_s < cfg["PASS_PRICE_LOW"] and comp_s < cfg["PASS_COMP_LOW"]:
        return opp, "PASS", round(conf, 2), \
            "Low value with high competition.", \
            "Monitor for price/competition changes.", subs

    return opp, "RESEARCH", round(conf, 2), \
        "Moderate signals — needs more data to classify.", \
        "Gather additional specs and pricing data; re-run scoring.", subs


# ──────────────────────────────────────────────
# SECTION D: Fingerprint functions (from engineering_fingerprint_v1.py)
# ──────────────────────────────────────────────

RE_NIIN = re.compile(r"\b\d{9}\b")
RE_NSN_DASH = re.compile(r"\b\d{4}-\d{2}-\d{3}-\d{4}\b")
RE_NSN_13 = re.compile(r"\b\d{13}\b")
RE_PN = re.compile(r"\b[A-Z0-9][A-Z0-9\-._/]{2,}\b", re.I)

# For FLIS import
RE_NIIN_EXACT = re.compile(r"^\d{9}$")
RE_NSN_DIGITS_EXACT = re.compile(r"^\d{13}$")
RE_NSN_DASH_EXACT = re.compile(r"^\d{4}-\d{2}-\d{3}-\d{4}$")
RE_CAGE_EXACT = re.compile(r"^[A-Z0-9]{5}$", re.I)
RE_PN_EXACT = re.compile(r"^[A-Z0-9][A-Z0-9\-._/]{2,}$", re.I)


def parse_rules(s: str) -> List[Tuple[str, str]]:
    out = []
    for part in (s or "").split(";"):
        part = part.strip()
        if not part or "=" not in part:
            continue
        k, v = part.split("=", 1)
        out.append((k.strip(), v.strip()))
    return out


def guess_doc_type(filename: str, rules: List[Tuple[str, str]]) -> str:
    f = filename.lower()
    for k, v in rules:
        if k.lower() in f:
            return v
    if f.endswith(".pdf") and ("ds" in f and "datasheet" in f):
        return "OEM_DATASHEET"
    if f.startswith("tm_") or "tm_" in f:
        return "MIL_TM"
    if f.startswith("spe") or "spe" in f:
        return "RFQ"
    return "PDF"


def extract_text(pdf_path: str, max_pages: int, max_chars: int) -> str:
    import pdfplumber
    text_parts = []
    with pdfplumber.open(pdf_path) as pdf:
        n = min(len(pdf.pages), max_pages)
        for i in range(n):
            t = pdf.pages[i].extract_text() or ""
            if t:
                text_parts.append(t)
            if sum(len(x) for x in text_parts) >= max_chars:
                break
    return "\n".join(text_parts)[:max_chars]


def near(text: str, idx: int, span: int = 60) -> str:
    a = max(0, idx - span)
    b = min(len(text), idx + span)
    return text[a:b]


def extract_key_facts(text: str, cfg: dict) -> dict:
    mats = [m.strip().lower() for m in (cfg.get("MATERIAL_KEYWORDS", "") or "").split(";") if m.strip()]
    valve_keys = [m.strip().lower() for m in (cfg.get("VALVE_TYPE_KEYWORDS", "") or "").split(";") if m.strip()]

    re_thread = re.compile(cfg.get("THREAD_REGEX", r""), re.I) if cfg.get("THREAD_REGEX") else None
    re_press = re.compile(cfg.get("PRESSURE_REGEX", r""), re.I) if cfg.get("PRESSURE_REGEX") else None
    re_temp = re.compile(cfg.get("TEMP_REGEX_F", r""), re.I) if cfg.get("TEMP_REGEX_F") else None
    re_std = re.compile(cfg.get("STD_REGEX", r""), re.I) if cfg.get("STD_REGEX") else None

    t_l = text.lower()
    found_mats = sorted({m for m in mats if m in t_l})
    found_types = sorted({k for k in valve_keys if k in t_l})
    threads = sorted({m.group(0).replace(" ", "") for m in re_thread.finditer(text)}) if re_thread else []
    standards = sorted({m.group(0).strip() for m in re_std.finditer(text)}) if re_std else []

    pressures = []
    if re_press:
        for m in re_press.finditer(text):
            try:
                val = float(m.group(1))
            except (ValueError, IndexError):
                continue
            ctx = near(text, m.start(), 80).lower()
            label = "pressure"
            if "proof" in ctx:
                label = "proof"
            elif "ultimate" in ctx or "burst" in ctx:
                label = "ultimate"
            elif "operat" in ctx or "working" in ctx:
                label = "operating"
            elif "leak" in ctx:
                label = "leak"
            pressures.append((label, val))

    p_best = {}
    for label, val in pressures:
        if label not in p_best:
            p_best[label] = val
        else:
            if label in ("proof", "ultimate", "operating", "pressure"):
                p_best[label] = max(p_best[label], val)
            else:
                p_best[label] = min(p_best[label], val)

    temps = []
    if re_temp:
        for m in re_temp.finditer(text):
            try:
                temps.append(float(m.group(1)))
            except (ValueError, TypeError, IndexError):
                pass

    ms = sorted(set(re.findall(r"\bMS\d{3,}(?:-[A-Z0-9]+)?\b", text, flags=re.I)))[:20]

    return {
        "materials": found_mats[:20], "types": found_types[:10],
        "threads": threads[:20], "standards": standards[:30],
        "pressure": p_best,
        "temp_f": {"min": min(temps) if temps else None, "max": max(temps) if temps else None},
        "ms_ids": ms,
    }


def format_key_facts(facts: dict, max_len: int = 300) -> str:
    bits = []
    if facts.get("types"):
        bits.append(f"type={facts['types'][0]}")
    if facts.get("threads"):
        bits.append(f"thread={facts['threads'][0]}")
    if facts.get("pressure"):
        if "operating" in facts["pressure"]:
            bits.append(f"op={facts['pressure']['operating']} psi")
        elif "pressure" in facts["pressure"]:
            bits.append(f"p={facts['pressure']['pressure']} psi")
        if "proof" in facts["pressure"]:
            bits.append(f"proof={facts['pressure']['proof']} psi")
        if "ultimate" in facts["pressure"]:
            bits.append(f"ult={facts['pressure']['ultimate']} psi")
    if facts.get("temp_f", {}).get("min") is not None:
        bits.append(f"T={facts['temp_f']['min']}..{facts['temp_f']['max']} F")
    if facts.get("ms_ids"):
        bits.append(f"MS={facts['ms_ids'][0]}")
    return "; ".join(bits)[:max_len]


def pn_candidates(text: str) -> List[str]:
    cands = []
    for m in RE_PN.finditer(text):
        s = m.group(0).strip()
        if len(s) < 4:
            continue
        if RE_NIIN.match(s) or RE_NSN_13.match(s) or RE_NSN_DASH.match(s):
            continue
        if s.lower().startswith("http"):
            continue
        if re.fullmatch(r"\d{1,6}", s):
            continue
        cands.append(s.upper())
        if len(cands) >= 40:
            break
    out, seen = [], set()
    for s in cands:
        if s not in seen:
            out.append(s)
            seen.add(s)
    return out[:25]


def find_ids(text: str) -> Tuple[Optional[str], Optional[str]]:
    niin, nsn = None, None
    m = RE_NIIN.search(text)
    if m:
        niin = m.group(0)
    m = RE_NSN_DASH.search(text)
    if m:
        nsn = m.group(0)
    else:
        m = RE_NSN_13.search(text)
        if m:
            nsn = m.group(0)
    return niin, nsn


def fp_confidence(doc_type: str, facts: dict, niin, nsn) -> float:
    base = {"OEM_DATASHEET": 90, "MIL_TM": 80, "RFQ": 65, "CATALOG": 60, "PDF": 50}.get(doc_type, 50)
    richness = 0
    if facts.get("threads"): richness += 6
    if facts.get("pressure"): richness += 8
    if facts.get("temp_f", {}).get("min") is not None: richness += 4
    if facts.get("standards"): richness += 4
    if facts.get("materials"): richness += 3
    if facts.get("types"): richness += 3
    if niin or nsn: richness += 5
    return min(100.0, base * 0.6 + richness * 4)


def merge_facts(existing: dict, new: dict, prefer_new: bool = False) -> dict:
    out = dict(existing)
    for k in ("materials", "types", "threads", "standards", "ms_ids"):
        out.setdefault(k, [])
        out[k] = sorted(set(out[k]) | set(new.get(k, [])))[:50]
    out.setdefault("pressure", {})
    for pk, pv in (new.get("pressure") or {}).items():
        if pk not in out["pressure"] or prefer_new:
            out["pressure"][pk] = pv
        else:
            if pk in ("proof", "ultimate", "operating", "pressure"):
                out["pressure"][pk] = max(out["pressure"][pk], pv)
            else:
                out["pressure"][pk] = min(out["pressure"][pk], pv)
    out.setdefault("temp_f", {"min": None, "max": None})
    nmin = (new.get("temp_f") or {}).get("min")
    nmax = (new.get("temp_f") or {}).get("max")
    if nmin is not None:
        out["temp_f"]["min"] = nmin if out["temp_f"]["min"] is None else min(out["temp_f"]["min"], nmin)
    if nmax is not None:
        out["temp_f"]["max"] = nmax if out["temp_f"]["max"] is None else max(out["temp_f"]["max"], nmax)
    return out


# ──────────────────────────────────────────────
# SECTION E: FLIS Import functions (from import_flis_reference_numbers_v7.py)
# ──────────────────────────────────────────────

def flis_score_column(values, header, detect_strict_cage, pn_minlen):
    n_nonblank = max(1, len([v for v in values if v]))
    vals = [v for v in values if v]

    def frac(pred):
        return sum(1 for v in vals if pred(v)) / n_nonblank

    header_l = (header or "").lower()

    f_niin = frac(lambda v: bool(RE_NIIN_EXACT.match(v.replace(" ", ""))))
    s_niin = 100.0 * f_niin

    def is_nsn(v):
        x = v.replace(" ", "")
        return bool(RE_NSN_DIGITS_EXACT.match(x)) or bool(RE_NSN_DASH_EXACT.match(x))
    f_nsn = frac(is_nsn)
    s_nsn = 100.0 * f_nsn

    def is_cage(v):
        x = v.strip().upper().replace(" ", "")
        if detect_strict_cage:
            return bool(RE_CAGE_EXACT.match(x))
        return bool(re.fullmatch(r"[A-Z0-9]{4,6}", x))
    f_cage = frac(is_cage)
    s_cage = 100.0 * f_cage

    def is_pn(v):
        x = v.strip().upper()
        if len(x) < pn_minlen:
            return False
        if RE_NIIN_EXACT.match(x) or RE_NSN_DIGITS_EXACT.match(x):
            return False
        if re.fullmatch(r"\d{1,6}", x):
            return False
        return bool(RE_PN_EXACT.match(x))
    f_pn = frac(is_pn)
    s_pn = 100.0 * f_pn

    def is_name(v):
        x = v.strip()
        if len(x) < 4:
            return False
        letters = sum(ch.isalpha() for ch in x)
        digits = sum(ch.isdigit() for ch in x)
        return letters >= 6 and digits <= 2 and "http" not in x.lower()
    f_name = frac(is_name)
    s_name = 100.0 * f_name

    def boost(base, kw_list):
        for kw in kw_list:
            if kw in header_l:
                return min(100.0, base + 15.0)
        return base

    s_niin = boost(s_niin, ["niin", "item ident"])
    s_nsn = boost(s_nsn, ["nsn", "national stock"])
    s_cage = boost(s_cage, ["cage", "fscm", "mfr code", "manufacturer code"])
    s_pn = boost(s_pn, ["pn", "p/n", "part number", "reference number", "ref no", "mfr pn"])
    s_name = boost(s_name, ["manufacturer", "mfr name", "name", "company"])

    return {"NIIN": s_niin, "NSN": s_nsn, "CAGE": s_cage, "OEM_PN": s_pn, "OEM_NAME": s_name}


def flis_pick_best(col_scores, target, used_cols):
    best, best_score = None, -1.0
    for col_idx, scores in col_scores:
        if col_idx in used_cols:
            continue
        sc = scores[target]
        if sc > best_score:
            best_score = sc
            best = col_idx
    if best is None:
        return None
    if best_score < 35.0 and target in ("NIIN", "NSN", "CAGE"):
        return None
    if best_score < 25.0 and target in ("OEM_PN", "OEM_NAME"):
        return None
    return best


def safe_str(v) -> str:
    return "" if v is None else str(v).strip()


def sample_xlsx(path, sheet, sample_rows):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        sheet = wb.sheetnames[0]
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    header = [safe_str(x) for x in next(it)]
    rows = []
    for i, r in enumerate(it, start=1):
        rows.append([safe_str(x) for x in r])
        if sample_rows and i >= sample_rows:
            break
    wb.close()
    return header, rows


def sample_csv(path, sample_rows):
    with open(path, "r", encoding="utf-8", errors="ignore", newline="") as f:
        reader = csv.reader(f)
        header = [safe_str(x) for x in next(reader)]
        rows = []
        for i, r in enumerate(reader, start=1):
            rows.append([safe_str(x) for x in r])
            if sample_rows and i >= sample_rows:
                break
    return header, rows


def flis_extract_fields(row, col_map):
    def get(key):
        idx = col_map.get(key)
        return row[idx] if idx is not None and idx < len(row) else ""
    return get("NIIN"), get("NSN"), get("CAGE"), get("OEM_PN"), get("OEM_NAME")


# ──────────────────────────────────────────────
# SECTION F: Pricing Import functions (from import_pricing_signals.py)
# ──────────────────────────────────────────────

def detect_pricing_columns(header_row):
    h = [str(x or "").strip().lower() for x in header_row]

    def find_any(keys):
        for k in keys:
            for i, name in enumerate(h):
                if k in name:
                    return i
        return None

    return {
        "niin": find_any(["niin", "national item identification"]),
        "nsn": find_any(["nsn", "national stock number"]),
        "price": find_any(["unit price", "unitprice", "price"]),
        "qty": find_any(["qty", "quantity"]),
        "date": find_any(["award date", "date"]),
        "supplier": find_any(["supplier", "vendor", "cage", "contractor"]),
    }


# ──────────────────────────────────────────────
# SECTION G: Pipeline Runners (background threads)
# ──────────────────────────────────────────────

def _update_log(log_id, status, rows=0, log_text=""):
    try:
        db = sqlite3.connect(str(DB_PATH), timeout=30)
        db.execute("PRAGMA journal_mode=WAL")
        db.execute(
            "UPDATE pipeline_log SET finished_at=datetime('now'), rows_processed=?, status=?, log_text=? WHERE id=?",
            (rows, status, str(log_text)[:2000], log_id)
        )
        db.commit()
        db.close()
    except Exception as e:
        log.error("Failed to update pipeline log %d: %s", log_id, str(e)[:200])


def run_decision_engine_worker(log_id):
    try:
        db = _worker_db()
        cfg = load_config_from_db(db, "decision", DEFAULTS)

        # Get all target NIINs from forecast (primary) or niin_targets (fallback)
        forecast_niins = db.execute("SELECT niin FROM forecast").fetchall()
        if forecast_niins:
            all_niins = [r["niin"] for r in forecast_niins]
        else:
            all_niins = [r["niin"] for r in db.execute("SELECT niin FROM niin_targets").fetchall()]

        if not all_niins:
            _update_log(log_id, "success", 0, "No NIINs to process. Import forecast or targets first.")
            db.close()
            return

        # Pre-load aggregate data in bulk for performance
        forecast_map = {}
        for r in db.execute("SELECT niin, total_demand, avg_monthly FROM forecast").fetchall():
            forecast_map[r["niin"]] = dict(r)

        char_counts = {}
        for r in db.execute("SELECT niin, COUNT(*) cnt FROM publog_characteristics GROUP BY niin").fetchall():
            char_counts[r["niin"]] = r["cnt"]

        ref_counts = {}
        cage_counts = {}
        for r in db.execute("SELECT niin, COUNT(*) cnt, COUNT(DISTINCT cage_code) cages FROM publog_references GROUP BY niin").fetchall():
            ref_counts[r["niin"]] = r["cnt"]
            cage_counts[r["niin"]] = r["cages"]

        mgmt_prices = {}
        for r in db.execute("""
            SELECT niin, unit_price FROM publog_management
            WHERE id IN (SELECT MAX(id) FROM publog_management GROUP BY niin)
        """).fetchall():
            mgmt_prices[r["niin"]] = r["unit_price"]

        ident_map = {}
        for r in db.execute("SELECT niin, item_name, crit_cd, dmil FROM publog_identification").fetchall():
            ident_map[r["niin"]] = dict(r)

        pricing_map = {}
        for r in db.execute("SELECT niin, median_unit_price, supplier_count FROM pricing_signals").fetchall():
            pricing_map[r["niin"]] = dict(r)

        nobid_map = {}
        for r in db.execute("SELECT niin, nobid_count FROM nobid_signals").fetchall():
            nobid_map[r["niin"]] = r["nobid_count"]

        count = 0
        for niin in all_niins:
            fc = forecast_map.get(niin, {})
            total_demand = fc.get("total_demand", 0) or 0
            char_count = char_counts.get(niin, 0)
            ref_count = ref_counts.get(niin, 0)
            cage_count = cage_counts.get(niin, 0)
            ident = ident_map.get(niin, {})
            item_name = ident.get("item_name", "")
            crit_cd = ident.get("crit_cd", "")
            dmil_val = ident.get("dmil", "")
            ps = pricing_map.get(niin, {})
            govt_price = mgmt_prices.get(niin)
            award_price = ps.get("median_unit_price")
            award_suppliers = ps.get("supplier_count", 0) or 0
            nobid_count = nobid_map.get(niin, 0)

            # Use govt price first, fall back to award price
            best_price = govt_price
            if best_price is None:
                best_price = award_price

            # Compute sub-scores
            demand_s = calc_demand_score(total_demand)
            char_s = calc_char_score(char_count)
            supply_s = calc_supply_score(cage_count, ref_count)
            price_s = calc_price_score(best_price)
            comp_s = calc_competition_score(award_suppliers, nobid_count)
            cots_r = calc_cots_readiness(char_s, supply_s, dmil_val, crit_cd, cfg)

            has_forecast = 1.0 if total_demand > 0 else 0.0
            has_chars = 1.0 if char_count > 0 else 0.0
            has_pricing = 1.0 if award_price is not None else 0.0
            has_refs = 1.0 if ref_count > 0 else 0.0
            has_mgmt = 1.0 if govt_price is not None else 0.0

            opp, action, conf, reason, nxt, subs = decide(
                demand_s, char_s, supply_s, price_s, comp_s, cots_r,
                has_forecast, has_chars, has_pricing, has_refs, has_mgmt, cfg
            )

            db.execute("""
                INSERT OR REPLACE INTO command_center
                (niin, opportunity_score, action, confidence, primary_reason, next_steps,
                 demand_score, char_score, supply_score, price_score, competition_score,
                 cots_readiness, total_demand, char_count, ref_count, cage_count,
                 govt_unit_price, item_name, crit_cd, dmil,
                 evidence_score, fp_score, docs_score, profit_score_val,
                 median_unit_price, supplier_count, updated_at)
                VALUES (?,?,?,?,?,?, ?,?,?,?,?, ?,?,?,?,?, ?,?,?,?, ?,?,?,?, ?,?,datetime('now'))
            """, (niin, opp, action, conf, reason, nxt,
                  subs["demand"], subs["chars"], subs["supply"], subs["price"], subs["competition"],
                  subs["cots_readiness"], total_demand, char_count, ref_count, cage_count,
                  best_price, item_name, crit_cd, dmil_val,
                  0, 0, 0, 0,  # legacy scores kept at 0
                  award_price, award_suppliers))
            count += 1

        db.commit()
        db.close()
        # Summary stats
        action_counts = {}
        for niin in all_niins:
            # recount from what we just wrote is expensive, use a simpler approach
            pass
        _update_log(log_id, "success", count, f"Scored {count} NIINs using forecast+PUBLOG data.")
        log.info("Decision engine completed: %d rows", count)
    except Exception as e:
        log.exception("Decision engine failed")
        _update_log(log_id, "error", log_text=str(e))


def run_import_pricing_worker(log_id, file_path, sheet_name, col_map_override=None):
    try:
        import openpyxl
        db = _worker_db()

        # load targets — use forecast NIINs first, then niin_targets, then None to accept all
        target_rows = db.execute("SELECT niin FROM forecast").fetchall()
        if not target_rows:
            target_rows = db.execute("SELECT niin FROM niin_targets").fetchall()
        targets = {r["niin"] for r in target_rows} if target_rows else None
        log.info("Pricing import: %d target NIINs loaded (%s)",
                 len(targets) if targets else 0, "filtering" if targets else "accepting all")

        pwb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name not in pwb.sheetnames:
            sheet_name = pwb.sheetnames[0]
        ws = pwb[sheet_name]

        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        colmap = detect_pricing_columns(header)
        if col_map_override:
            colmap.update(col_map_override)

        if colmap.get("niin") is None:
            raise ValueError("Could not detect NIIN column in pricing file")

        agg = {}
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_count += 1
            n = norm_niin(row[colmap["niin"]] if colmap["niin"] is not None else None)
            if not n:
                continue
            if targets is not None and n not in targets:
                continue

            price = row[colmap["price"]] if colmap.get("price") is not None else None
            qty = row[colmap["qty"]] if colmap.get("qty") is not None else None
            dt = row[colmap["date"]] if colmap.get("date") is not None else None
            sup = row[colmap["supplier"]] if colmap.get("supplier") is not None else None

            try:
                price_f = float(price) if price not in (None, "") else None
            except (ValueError, TypeError):
                price_f = None
            try:
                qty_i = int(float(qty)) if qty not in (None, "") else 0
            except (ValueError, TypeError):
                qty_i = 0

            rec = agg.setdefault(n, {"prices": [], "total_qty": 0, "total_awards": 0,
                                     "last_date": None, "last_price": None, "suppliers": {}})
            if price_f is not None:
                rec["prices"].append(price_f)
            rec["total_qty"] += qty_i
            rec["total_awards"] += 1

            if dt:
                try:
                    dtx = dt if isinstance(dt, datetime) else datetime.fromisoformat(str(dt)[:10])
                    if rec["last_date"] is None or dtx > rec["last_date"]:
                        rec["last_date"] = dtx
                        if price_f is not None:
                            rec["last_price"] = price_f
                except (ValueError, TypeError, AttributeError):
                    pass

            if sup:
                s = str(sup).strip()
                rec["suppliers"][s] = rec["suppliers"].get(s, 0) + 1

        pwb.close()

        # write to DB
        db.execute("DELETE FROM pricing_signals")
        src_name = Path(file_path).name
        for n, rec in agg.items():
            prices = rec["prices"]
            if prices:
                last_price = rec["last_price"] if rec["last_price"] is not None else prices[-1]
                med_price = statistics.median(prices)
                min_p, max_p = min(prices), max(prices)
            else:
                last_price = med_price = min_p = max_p = None

            suppliers = rec["suppliers"]
            sup_count = len(suppliers)
            top_sup = max(suppliers.items(), key=lambda kv: kv[1])[0] if suppliers else ""
            last_dt = rec["last_date"].isoformat() if rec["last_date"] else ""

            db.execute("INSERT OR IGNORE INTO niins (niin) VALUES (?)", (n,))
            db.execute("""
                INSERT OR REPLACE INTO pricing_signals
                (niin, last_unit_price, median_unit_price, min_unit_price, max_unit_price,
                 total_awards, total_qty, last_award_date, supplier_count, top_supplier, source_file)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """, (n, last_price, med_price, min_p, max_p,
                  rec["total_awards"], rec["total_qty"], last_dt, sup_count, top_sup, src_name))

            # Also update command_center if row exists
            db.execute("""
                INSERT INTO command_center (niin, median_unit_price, supplier_count)
                VALUES (?, ?, ?)
                ON CONFLICT(niin) DO UPDATE SET
                    median_unit_price=excluded.median_unit_price,
                    supplier_count=excluded.supplier_count,
                    updated_at=datetime('now')
            """, (n, med_price, sup_count))

        db.commit()
        db.close()
        skip_info = f" (targets={len(targets)})" if targets else " (no filter)"
        _update_log(log_id, "success", len(agg),
                    f"Processed {row_count} source rows → {len(agg)} NIINs with pricing{skip_info}")
        log.info("Pricing import completed: %d NIINs from %d rows", len(agg), row_count)
    except Exception as e:
        log.exception("Pricing import failed")
        _update_log(log_id, "error", log_text=str(e))


def run_fingerprint_worker(log_id):
    try:
        import pdfplumber  # noqa
        db = _worker_db()
        cfg = load_config_from_db(db, "fingerprint", FINGERPRINT_DEFAULTS)

        pdf_dir = cfg.get("PDF_DIR", "").strip() or str(PDF_DIR)
        max_pages = int(cfg.get("MAX_PAGES_PER_PDF", "12") or "12")
        max_chars = int(cfg.get("MAX_CHARS_PER_PDF", "180000") or "180000")
        rules = parse_rules(cfg.get("DOC_TYPE_RULES", ""))

        pdfs = sorted(str(p) for p in Path(pdf_dir).rglob("*.pdf"))
        if not pdfs:
            _update_log(log_id, "success", 0, f"No PDFs found in {pdf_dir}")
            db.close()
            return

        db.execute("DELETE FROM doc_extract")
        by_niin = defaultdict(lambda: {"fingerprint": {}, "evidence_docs": [], "conf": 0.0, "nsn": None})
        failed = 0

        for i, p in enumerate(pdfs, start=1):
            fn = os.path.basename(p)
            dt = guess_doc_type(fn, rules)
            doc_id = f"D{i:04d}"

            try:
                text = extract_text(p, max_pages, max_chars)
            except Exception as e:
                db.execute("""
                    INSERT INTO doc_extract (doc_id, filename, doc_type, confidence, notes)
                    VALUES (?,?,?,0,?)
                """, (doc_id, fn, dt, f"FAILED: {e}"))
                failed += 1
                continue

            niin, nsn = find_ids(text)
            cands = pn_candidates(text)
            facts = extract_key_facts(text, cfg)
            conf = fp_confidence(dt, facts, niin, nsn)
            kf = format_key_facts(facts, 240)

            extract = {"doc_type": dt, "ids": {"niin": niin, "nsn": nsn},
                       "pn_candidates": cands, "facts": facts}

            db.execute("""
                INSERT INTO doc_extract
                (doc_id, filename, doc_type, niin, nsn, pn_candidates, key_facts, extract_json, confidence, notes)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (doc_id, fn, dt, niin or "", nsn or "", ", ".join(cands[:8]),
                  kf, json.dumps(extract, ensure_ascii=False)[:32000], round(conf, 1),
                  f"pages<={max_pages}"))

            if niin:
                db.execute("INSERT OR IGNORE INTO niins (niin) VALUES (?)", (niin,))
                agg = by_niin[niin]
                prefer = (dt == "OEM_DATASHEET")
                agg["fingerprint"] = merge_facts(agg["fingerprint"], facts, prefer_new=prefer)
                agg["evidence_docs"].append(fn)
                agg["conf"] = max(agg["conf"], conf)
                if nsn and not agg["nsn"]:
                    agg["nsn"] = nsn

        # write fingerprint_by_niin
        db.execute("DELETE FROM fingerprint_by_niin")
        for niin, v in by_niin.items():
            db.execute("""
                INSERT INTO fingerprint_by_niin (niin, nsn, fingerprint_json, evidence_docs, confidence, top_facts)
                VALUES (?,?,?,?,?,?)
            """, (niin, v.get("nsn") or "", json.dumps(v["fingerprint"], ensure_ascii=False)[:32000],
                  ", ".join(v["evidence_docs"][:10]), round(v["conf"], 1),
                  format_key_facts(v["fingerprint"], 300)))

            # update command_center
            db.execute("""
                INSERT INTO command_center (niin, fp_conf, docs_total)
                VALUES (?, ?, ?)
                ON CONFLICT(niin) DO UPDATE SET
                    fp_conf=excluded.fp_conf,
                    docs_total=excluded.docs_total,
                    updated_at=datetime('now')
            """, (niin, round(v["conf"], 1), len(v["evidence_docs"])))

        db.commit()
        db.close()
        _update_log(log_id, "success", len(pdfs),
                    f"Scanned {len(pdfs)} PDFs (failed={failed}); {len(by_niin)} NIIN fingerprints")
        log.info("Fingerprint completed: %d PDFs, %d NIINs", len(pdfs), len(by_niin))
    except Exception as e:
        log.exception("Fingerprint failed")
        _update_log(log_id, "error", log_text=str(e))


def run_flis_import_worker(log_id, file_path, file_type, sheet_name):
    try:
        db = _worker_db()

        sample_rows_count = 200
        strict_cage = False
        pn_minlen = 3

        if file_type == "xlsx":
            header, rows = sample_xlsx(file_path, sheet_name, sample_rows_count)
        else:
            header, rows = sample_csv(file_path, sample_rows_count)

        ncols = len(header)
        col_samples = []
        for c in range(ncols):
            vals = []
            for r in rows:
                if c < len(r) and r[c]:
                    vals.append(r[c])
                if len(vals) >= 30:
                    break
            col_samples.append(vals)

        col_scores = []
        for c in range(ncols):
            scores = flis_score_column(col_samples[c], header[c] if c < len(header) else "", strict_cage, pn_minlen)
            col_scores.append((c, scores))

        # write detection report
        db.execute("DELETE FROM flis_detected_schema")
        src_name = os.path.basename(file_path)
        for c, scores in col_scores:
            best_guess = max(scores.items(), key=lambda kv: kv[1])[0]
            best_score = scores[best_guess]
            samp = ", ".join(col_samples[c][:8])
            db.execute("""
                INSERT INTO flis_detected_schema
                (source_file, sheet_name, col_index, col_letter, header_name, best_guess, score, explanation, sample_values, recommended_key)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (src_name, sheet_name if file_type == "xlsx" else "", c + 1,
                  chr(65 + c) if c < 26 else f"C{c+1}", header[c] if c < len(header) else "",
                  best_guess, round(best_score, 1), f"{best_score:.0f}% match",
                  samp[:800], f"COL_{best_guess}"))

        # auto pick columns
        col_map = {}
        used = set()
        for tgt in ("NIIN", "NSN", "CAGE", "OEM_PN", "OEM_NAME"):
            pick = flis_pick_best(col_scores, tgt, used)
            if pick is not None:
                col_map[tgt] = pick
                used.add(pick)

        # import rows
        db.execute("DELETE FROM flis_refnums")
        ts = now_iso()
        count = 0

        if file_type == "xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
            it = ws.iter_rows(values_only=True)
            next(it)  # skip header
            for row_num, r in enumerate(it, start=2):
                r = [safe_str(x) for x in r]
                niin, nsn, cage, pn, name = flis_extract_fields(r, col_map)
                if not pn and not niin and not nsn:
                    continue
                db.execute("""
                    INSERT INTO flis_refnums (niin, nsn, cage, oem_pn, oem_name, source_file, source_row, imported_at)
                    VALUES (?,?,?,?,?,?,?,?)
                """, (niin, nsn, cage, pn, name, src_name, row_num, ts))
                if niin and len(niin) == 9:
                    db.execute("INSERT OR IGNORE INTO niins (niin) VALUES (?)", (niin,))
                count += 1
            wb.close()
        else:
            with open(file_path, "r", encoding="utf-8", errors="ignore", newline="") as f:
                reader = csv.reader(f)
                next(reader)
                for row_num, r in enumerate(reader, start=2):
                    r = [safe_str(x) for x in r]
                    niin, nsn, cage, pn, name = flis_extract_fields(r, col_map)
                    if not pn and not niin and not nsn:
                        continue
                    db.execute("""
                        INSERT INTO flis_refnums (niin, nsn, cage, oem_pn, oem_name, source_file, source_row, imported_at)
                        VALUES (?,?,?,?,?,?,?,?)
                    """, (niin, nsn, cage, pn, name, src_name, row_num, ts))
                    if niin and len(niin) == 9:
                        db.execute("INSERT OR IGNORE INTO niins (niin) VALUES (?)", (niin,))
                    count += 1

        db.commit()
        db.close()
        _update_log(log_id, "success", count, f"Imported {count} rows, detected {len(col_map)} columns")
        log.info("FLIS import completed: %d rows", count)
    except Exception as e:
        log.exception("FLIS import failed")
        _update_log(log_id, "error", log_text=str(e))


def run_import_nobid_worker(log_id, file_path):
    try:
        import openpyxl
        db = _worker_db()

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header = [str(x or "").strip().lower() for x in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

        def find_col(keys):
            for k in keys:
                for i, h in enumerate(header):
                    if k in h:
                        return i
            return None

        c_niin = find_col(["niin", "national item"])
        c_count = find_col(["count", "no bid count", "nobid count"])
        c_last = find_col(["last close", "close date", "last_close"])
        c_value = find_col(["pr value", "total value", "total_pr"])
        c_fsc = find_col(["fsc", "federal supply"])
        c_nomen = find_col(["nomenclature", "item name", "description"])

        if c_niin is None:
            raise ValueError("Could not detect NIIN column in NoBid file")

        db.execute("DELETE FROM nobid_signals")
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            n = norm_niin(row[c_niin] if c_niin is not None else None)
            if not n:
                continue
            nb_count = to_int(row[c_count]) if c_count is not None else 0
            last_close = str(row[c_last] or "") if c_last is not None else ""
            total_val = to_float(row[c_value]) if c_value is not None else 0
            fsc = str(row[c_fsc] or "") if c_fsc is not None else ""
            nomen = str(row[c_nomen] or "") if c_nomen is not None else ""

            db.execute("INSERT OR IGNORE INTO niins (niin) VALUES (?)", (n,))
            db.execute("""
                INSERT OR REPLACE INTO nobid_signals
                (niin, nobid_count, last_close, total_pr_value, fsc, nomenclature)
                VALUES (?,?,?,?,?,?)
            """, (n, nb_count, last_close, total_val or 0, fsc, nomen))

            # update command_center nobid_score
            # simple scoring: clamp(count * 15 + recency bonus)
            score = clamp(nb_count * 15)
            db.execute("""
                INSERT INTO command_center (niin, nobid_score)
                VALUES (?, ?)
                ON CONFLICT(niin) DO UPDATE SET
                    nobid_score=excluded.nobid_score,
                    updated_at=datetime('now')
            """, (n, score))
            count += 1

        wb.close()
        db.commit()
        db.close()
        _update_log(log_id, "success", count)
        log.info("NoBid import completed: %d NIINs", count)
    except Exception as e:
        log.exception("NoBid import failed")
        _update_log(log_id, "error", log_text=str(e))


def run_excel_migration_worker(log_id, file_path):
    try:
        import openpyxl
        db = _worker_db()
        wb = openpyxl.load_workbook(file_path, data_only=True)
        migrated = []

        # Migrate NIIN_Target
        if "NIIN_Target" in wb.sheetnames:
            ws = wb["NIIN_Target"]
            cnt = 0
            for row in ws.iter_rows(min_row=1, values_only=True):
                n = norm_niin(row[0] if row else None)
                if n:
                    db.execute("INSERT OR IGNORE INTO niins (niin) VALUES (?)", (n,))
                    db.execute("INSERT OR IGNORE INTO niin_targets (niin) VALUES (?)", (n,))
                    cnt += 1
            migrated.append(f"NIIN_Target: {cnt}")

        # Migrate NSN_Command_Center (HDR=6)
        if "NSN_Command_Center" in wb.sheetnames:
            ws = wb["NSN_Command_Center"]
            HDR = 6
            headers = {}
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(HDR, c).value or "").strip()
                if val:
                    headers[val] = c

            def gc(name):
                return headers.get(name)

            cnt = 0
            for r in range(HDR + 1, ws.max_row + 1):
                niin_val = ws.cell(r, gc("NIIN")).value if gc("NIIN") else None
                n = norm_niin(niin_val)
                if not n:
                    continue

                def cell(col_name):
                    c = gc(col_name)
                    return ws.cell(r, c).value if c else None

                db.execute("INSERT OR IGNORE INTO niins (niin) VALUES (?)", (n,))
                db.execute("""
                    INSERT OR REPLACE INTO command_center
                    (niin, coverage, docs_total, fp_conf, nobid_score, opportunity_score,
                     action, confidence, primary_reason, next_steps,
                     best_url, top_gaps, pn_dup_count, median_unit_price, supplier_count)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    n,
                    to_float(cell("Coverage (0-100)")) or 0,
                    to_int(cell("Docs Total")),
                    to_float(cell("FP Conf (0-100)")) or 0,
                    to_float(cell("NoBid Score (0-100)")) or 0,
                    to_float(cell("Opportunity Score (0-100)")) or 0,
                    str(cell("Action v3") or cell("Action v2") or cell("Action v1") or ""),
                    to_float(cell("Confidence v3 (0-1)") or cell("Confidence v2 (0-1)") or cell("Confidence v1 (0-1)")) or 0,
                    str(cell("Primary Reason") or ""),
                    str(cell("Next Steps") or ""),
                    str(cell("Best URL") or ""),
                    str(cell("Top Gaps") or ""),
                    to_int(cell("PN Dup Count")),
                    to_float(cell("Median Unit Price")),
                    to_int(cell("Supplier Count")),
                ))
                cnt += 1
            migrated.append(f"Command_Center: {cnt}")

        # Migrate NoBid_Signals
        if "NoBid_Signals" in wb.sheetnames:
            ws = wb["NoBid_Signals"]
            cnt = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                n = norm_niin(row[0])
                if not n:
                    continue
                db.execute("INSERT OR IGNORE INTO niins (niin) VALUES (?)", (n,))
                db.execute("""
                    INSERT OR REPLACE INTO nobid_signals (niin, nobid_count, last_close, total_pr_value, fsc, nomenclature)
                    VALUES (?,?,?,?,?,?)
                """, (n, to_int(row[1] if len(row) > 1 else 0),
                      str(row[2] or "") if len(row) > 2 else "",
                      to_float(row[4]) or 0 if len(row) > 4 else 0,
                      str(row[5] or "") if len(row) > 5 else "",
                      str(row[6] or "") if len(row) > 6 else ""))
                cnt += 1
            migrated.append(f"NoBid_Signals: {cnt}")

        # Migrate Pricing_Signals
        if "Pricing_Signals" in wb.sheetnames:
            ws = wb["Pricing_Signals"]
            cnt = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                n = norm_niin(row[0])
                if not n:
                    continue
                db.execute("INSERT OR IGNORE INTO niins (niin) VALUES (?)", (n,))
                db.execute("""
                    INSERT OR REPLACE INTO pricing_signals
                    (niin, last_unit_price, median_unit_price, min_unit_price, max_unit_price,
                     total_awards, total_qty, last_award_date, supplier_count, top_supplier, source_file)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """, (n,
                      to_float(row[1]) if len(row) > 1 else None,
                      to_float(row[2]) if len(row) > 2 else None,
                      to_float(row[3]) if len(row) > 3 else None,
                      to_float(row[4]) if len(row) > 4 else None,
                      to_int(row[5]) if len(row) > 5 else 0,
                      to_int(row[6]) if len(row) > 6 else 0,
                      str(row[7] or "") if len(row) > 7 else "",
                      to_int(row[8]) if len(row) > 8 else 0,
                      str(row[9] or "") if len(row) > 9 else "",
                      str(row[10] or "") if len(row) > 10 else ""))
                cnt += 1
            migrated.append(f"Pricing_Signals: {cnt}")

        # Migrate config sheets
        for sheet_name, section, desc_map in [
            ("Decision_Config", "decision", DECISION_DESCRIPTIONS),
            ("Fingerprint_Config", "fingerprint", FINGERPRINT_DESCRIPTIONS),
        ]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                cnt = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        k = str(row[0]).strip()
                        v = "" if row[1] is None else str(row[1]).strip()
                        db.execute("""
                            INSERT OR REPLACE INTO config (key, value, section, description)
                            VALUES (?,?,?,?)
                        """, (k, v, section, desc_map.get(k, "")))
                        cnt += 1
                migrated.append(f"{sheet_name}: {cnt}")

        wb.close()
        db.commit()
        db.close()
        summary = "; ".join(migrated)
        _update_log(log_id, "success", log_text=summary)
        log.info("Excel migration completed: %s", summary)
    except Exception as e:
        log.exception("Excel migration failed")
        _update_log(log_id, "error", log_text=str(e))


def run_publog_import_worker(log_id, publog_dir, min_demand=0):
    """Import PUBLOG data from ZIPs for target NIINs only (memory-efficient streaming)."""
    import zipfile
    try:
        db = _worker_db()
        publog = Path(publog_dir)

        # Gather target NIINs — filter by min_demand if specified
        if min_demand > 0:
            target_rows = db.execute(
                "SELECT niin FROM forecast WHERE total_demand >= ?", (min_demand,)
            ).fetchall()
            targets = {r["niin"] for r in target_rows}
            log.info("PUBLOG import: filtered to %d NIINs with demand >= %d", len(targets), min_demand)
        else:
            target_rows = db.execute("SELECT niin FROM niin_targets").fetchall()
            if target_rows:
                targets = {r["niin"] for r in target_rows}
            else:
                target_rows = db.execute("SELECT niin FROM command_center").fetchall()
                targets = {r["niin"] for r in target_rows}

        if not targets:
            _update_log(log_id, "error", log_text="No target NIINs found. Import NIINs first via Excel migration or manual entry.")
            db.close()
            return

        log.info("PUBLOG import: %d target NIINs", len(targets))
        msgs = []

        # ── 1. IDENTIFICATION.zip → P_FLIS_NSN.CSV ──
        ident_zip = publog / "IDENTIFICATION.zip"
        if ident_zip.exists():
            cnt = 0
            db.execute("DELETE FROM publog_identification")
            with zipfile.ZipFile(str(ident_zip), 'r') as zf:
                csv_name = None
                for n in zf.namelist():
                    if 'P_FLIS_NSN' in n.upper() and n.upper().endswith('.CSV'):
                        csv_name = n
                        break
                if csv_name:
                    with zf.open(csv_name) as f:
                        reader = csv.reader(io.TextIOWrapper(f, encoding='utf-8', errors='ignore'))
                        header = next(reader, None)
                        for row in reader:
                            if len(row) < 4:
                                continue
                            niin = row[1].strip() if len(row) > 1 else ""
                            if niin not in targets:
                                continue
                            fsc = row[0].strip() if len(row) > 0 else ""
                            inc = row[2].strip() if len(row) > 2 else ""
                            item_name = row[3].strip() if len(row) > 3 else ""
                            sos = row[4].strip() if len(row) > 4 else ""
                            end_item = row[5].strip() if len(row) > 5 else ""
                            cancelled = row[6].strip() if len(row) > 6 else ""
                            db.execute("""
                                INSERT OR REPLACE INTO publog_identification
                                (niin, fsc, inc, item_name, sos, end_item_name, cancelled_niin)
                                VALUES (?,?,?,?,?,?,?)
                            """, (niin, fsc, inc, item_name, sos, end_item, cancelled))
                            cnt += 1
                # Also try V_FLIS_IDENTIFICATION for extended fields
                vid_name = None
                for n in zf.namelist():
                    if 'V_FLIS_IDENTIFICATION' in n.upper() and n.upper().endswith('.CSV'):
                        vid_name = n
                        break
                if vid_name:
                    with zf.open(vid_name) as f:
                        reader = csv.reader(io.TextIOWrapper(f, encoding='utf-8', errors='ignore'))
                        vid_header = next(reader, [])
                        # Build column map from actual headers
                        vid_cols = {h.strip().upper(): i for i, h in enumerate(vid_header)}
                        for row in reader:
                            if len(row) < 2:
                                continue
                            niin = row[vid_cols.get("NIIN", 0)].strip()
                            if niin not in targets:
                                continue
                            def _vc(col):
                                idx = vid_cols.get(col)
                                return row[idx].strip() if idx is not None and idx < len(row) else ""
                            fiig = _vc("FIIG")
                            crit_cd = _vc("CRIT_CD")
                            dmil = _vc("DMIL")
                            niin_asgmt = _vc("NIIN_ASGMT")
                            pmic = _vc("PMIC")
                            hmic = _vc("HMIC")
                            db.execute("""
                                UPDATE publog_identification SET
                                    fiig=?, crit_cd=?, dmil=?, niin_asgmt=?, pmic=?, hmic=?
                                WHERE niin=?
                            """, (fiig, crit_cd, dmil, niin_asgmt, pmic, hmic, niin))
            db.commit()
            msgs.append(f"Identification: {cnt} NIINs")
            log.info("PUBLOG identification: %d rows", cnt)

        # ── 2. CHARACTERISTICS.zip → V_CHARACTERISTICS.CSV ──
        char_zip = publog / "CHARACTERISTICS.zip"
        if char_zip.exists():
            cnt = 0
            db.execute("DELETE FROM publog_characteristics")
            with zipfile.ZipFile(str(char_zip), 'r') as zf:
                csv_name = None
                for n in zf.namelist():
                    if 'CHARACTERISTICS' in n.upper() and n.upper().endswith('.CSV'):
                        csv_name = n
                        break
                if csv_name:
                    with zf.open(csv_name) as f:
                        reader = csv.reader(io.TextIOWrapper(f, encoding='utf-8', errors='ignore'))
                        next(reader, None)
                        for row in reader:
                            if len(row) < 2:
                                continue
                            niin = row[0].strip()
                            if niin not in targets:
                                continue
                            mrc = row[1].strip() if len(row) > 1 else ""
                            req_stmt = row[2].strip() if len(row) > 2 else ""
                            clear_text = row[3].strip() if len(row) > 3 else ""
                            db.execute("""
                                INSERT INTO publog_characteristics (niin, mrc, requirements_statement, clear_text_reply)
                                VALUES (?,?,?,?)
                            """, (niin, mrc, req_stmt, clear_text))
                            cnt += 1
            db.commit()
            msgs.append(f"Characteristics: {cnt} rows")
            log.info("PUBLOG characteristics: %d rows", cnt)

        # ── 3. MANAGEMENT.zip → V_FLIS_MANAGEMENT.CSV ──
        mgmt_zip = publog / "MANAGEMENT.zip"
        if mgmt_zip.exists():
            cnt = 0
            db.execute("DELETE FROM publog_management")
            with zipfile.ZipFile(str(mgmt_zip), 'r') as zf:
                csv_name = None
                for n in zf.namelist():
                    if 'V_FLIS_MANAGEMENT' in n.upper() and n.upper().endswith('.CSV') and 'FUTURE' not in n.upper():
                        csv_name = n
                        break
                if csv_name:
                    with zf.open(csv_name) as f:
                        reader = csv.reader(io.TextIOWrapper(f, encoding='utf-8', errors='ignore'))
                        next(reader, None)
                        for row in reader:
                            if len(row) < 2:
                                continue
                            niin = row[0].strip()
                            if niin not in targets:
                                continue
                            db.execute("""
                                INSERT INTO publog_management
                                (niin, effective_date, moe, aac, sos, ui, ui_conv_fac, unit_price, qup, ciic, slc, usc)
                                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                            """, (
                                niin,
                                row[1].strip() if len(row) > 1 else "",
                                row[2].strip() if len(row) > 2 else "",
                                row[3].strip() if len(row) > 3 else "",
                                row[4].strip() if len(row) > 4 else "",
                                row[6].strip() if len(row) > 6 else "",
                                row[7].strip() if len(row) > 7 else "",
                                row[8].strip() if len(row) > 8 else "",
                                row[9].strip() if len(row) > 9 else "",
                                row[10].strip() if len(row) > 10 else "",
                                row[11].strip() if len(row) > 11 else "",
                                row[15].strip() if len(row) > 15 else "",
                            ))
                            cnt += 1
            db.commit()
            msgs.append(f"Management: {cnt} rows")
            log.info("PUBLOG management: %d rows", cnt)

        # ── 4. REFERENCE.zip → V_FLIS_PART.CSV ──
        ref_zip = publog / "REFERENCE.zip"
        if ref_zip.exists():
            cnt = 0
            cage_codes_needed = set()
            db.execute("DELETE FROM publog_references")
            with zipfile.ZipFile(str(ref_zip), 'r') as zf:
                csv_name = None
                for n in zf.namelist():
                    if 'V_FLIS_PART' in n.upper() and n.upper().endswith('.CSV'):
                        csv_name = n
                        break
                if csv_name:
                    with zf.open(csv_name) as f:
                        reader = csv.reader(io.TextIOWrapper(f, encoding='utf-8', errors='ignore'))
                        next(reader, None)
                        for row in reader:
                            if len(row) < 3:
                                continue
                            niin = row[0].strip()
                            if niin not in targets:
                                continue
                            part_number = row[1].strip() if len(row) > 1 else ""
                            cage_code = row[2].strip() if len(row) > 2 else ""
                            cage_status = row[3].strip() if len(row) > 3 else ""
                            rncc = row[4].strip() if len(row) > 4 else ""
                            rnvc = row[5].strip() if len(row) > 5 else ""
                            dac = row[6].strip() if len(row) > 6 else ""
                            rnaac = row[7].strip() if len(row) > 7 else ""
                            db.execute("""
                                INSERT INTO publog_references
                                (niin, part_number, cage_code, cage_status, rncc, rnvc, dac, rnaac)
                                VALUES (?,?,?,?,?,?,?,?)
                            """, (niin, part_number, cage_code, cage_status, rncc, rnvc, dac, rnaac))
                            if cage_code:
                                cage_codes_needed.add(cage_code)
                            cnt += 1
            db.commit()
            msgs.append(f"References: {cnt} rows ({len(cage_codes_needed)} unique CAGEs)")
            log.info("PUBLOG references: %d rows, %d CAGEs", cnt, len(cage_codes_needed))

            # ── 5. CAGE.zip → P_CAGE.CSV (only needed CAGEs) ──
            cage_zip = publog / "CAGE.zip"
            if cage_zip.exists() and cage_codes_needed:
                ccnt = 0
                db.execute("DELETE FROM publog_cage")
                with zipfile.ZipFile(str(cage_zip), 'r') as zf:
                    csv_name = None
                    for n in zf.namelist():
                        if 'P_CAGE' in n.upper() and n.upper().endswith('.CSV'):
                            csv_name = n
                            break
                    if csv_name:
                        with zf.open(csv_name) as f:
                            reader = csv.reader(io.TextIOWrapper(f, encoding='utf-8', errors='ignore'))
                            next(reader, None)
                            for row in reader:
                                if len(row) < 5:
                                    continue
                                cage = row[0].strip()
                                if cage not in cage_codes_needed:
                                    continue
                                db.execute("""
                                    INSERT OR REPLACE INTO publog_cage
                                    (cage_code, cage_status, type, company, city, state_province, zip_postal_zone, country)
                                    VALUES (?,?,?,?,?,?,?,?)
                                """, (
                                    cage,
                                    row[1].strip() if len(row) > 1 else "",
                                    row[2].strip() if len(row) > 2 else "",
                                    row[4].strip() if len(row) > 4 else "",
                                    row[5].strip() if len(row) > 5 else "",
                                    row[6].strip() if len(row) > 6 else "",
                                    row[7].strip() if len(row) > 7 else "",
                                    row[8].strip() if len(row) > 8 else "",
                                ))
                                ccnt += 1
                db.commit()
                msgs.append(f"CAGE: {ccnt} suppliers")
                log.info("PUBLOG CAGE: %d rows", ccnt)

        # ── 6. H-SERIES.zip → FSC + INC lookups ──
        hs_zip = publog / "H-SERIES.zip"
        if hs_zip.exists():
            # Collect FSC/INC values we need
            fsc_needed = set()
            inc_needed = set()
            for r in db.execute("SELECT fsc, inc FROM publog_identification").fetchall():
                if r["fsc"]:
                    fsc_needed.add(r["fsc"])
                if r["inc"]:
                    inc_needed.add(r["inc"])

            fcnt = icnt = 0
            db.execute("DELETE FROM publog_hseries_fsc")
            db.execute("DELETE FROM publog_hseries_inc")

            with zipfile.ZipFile(str(hs_zip), 'r') as zf:
                # FSC titles from P_H2_PICK
                for n in zf.namelist():
                    if 'P_H2_PICK' in n.upper() and n.upper().endswith('.CSV'):
                        with zf.open(n) as f:
                            reader = csv.reader(io.TextIOWrapper(f, encoding='utf-8', errors='ignore'))
                            next(reader, None)
                            for row in reader:
                                if len(row) < 4:
                                    continue
                                fsc = row[1].strip() if len(row) > 1 else ""
                                if fsc not in fsc_needed:
                                    continue
                                db.execute("""
                                    INSERT OR REPLACE INTO publog_hseries_fsc (fsc, fsc_title, fsg_title)
                                    VALUES (?,?,?)
                                """, (fsc, row[3].strip() if len(row) > 3 else "",
                                      row[2].strip() if len(row) > 2 else ""))
                                fcnt += 1
                        break

                # INC names from P_H6_PICK
                for n in zf.namelist():
                    if 'P_H6_PICK' in n.upper() and n.upper().endswith('.CSV'):
                        with zf.open(n) as f:
                            reader = csv.reader(io.TextIOWrapper(f, encoding='utf-8', errors='ignore'))
                            next(reader, None)
                            for row in reader:
                                if len(row) < 2:
                                    continue
                                inc = row[0].strip()
                                if inc not in inc_needed:
                                    continue
                                db.execute("""
                                    INSERT OR REPLACE INTO publog_hseries_inc (inc, item_name, fiig, fsc)
                                    VALUES (?,?,?,?)
                                """, (inc, row[1].strip() if len(row) > 1 else "",
                                      row[2].strip() if len(row) > 2 else "",
                                      row[4].strip() if len(row) > 4 else ""))
                                icnt += 1
                        break

                # INC definitions from V_H6_NAME_INC
                for n in zf.namelist():
                    if 'V_H6_NAME_INC' in n.upper() and n.upper().endswith('.CSV'):
                        with zf.open(n) as f:
                            reader = csv.reader(io.TextIOWrapper(f, encoding='utf-8', errors='ignore'))
                            next(reader, None)
                            for row in reader:
                                if len(row) < 2:
                                    continue
                                inc = row[0].strip()
                                if inc not in inc_needed:
                                    continue
                                definition = row[8].strip() if len(row) > 8 else ""
                                if definition:
                                    db.execute("""
                                        UPDATE publog_hseries_inc SET definition=? WHERE inc=?
                                    """, (definition, inc))
                        break

            db.commit()
            msgs.append(f"H-Series: {fcnt} FSCs, {icnt} INCs")
            log.info("PUBLOG H-Series: %d FSCs, %d INCs", fcnt, icnt)

        db.close()
        summary = "; ".join(msgs) if msgs else "No PUBLOG ZIPs found"
        _update_log(log_id, "success", log_text=summary)
        log.info("PUBLOG import completed: %s", summary)
    except Exception as e:
        log.exception("PUBLOG import failed")
        _update_log(log_id, "error", log_text=str(e))


def run_forecast_import_worker(log_id, file_path):
    """Import DIBBS SRVA forecast file (.txt pipe-delimited or .xls HTML table)."""
    try:
        db = _worker_db()
        fpath = Path(file_path)
        source_name = fpath.name

        rows_data = []

        if fpath.suffix.lower() == '.txt':
            # Pipe-delimited format (DIBBS: leading/trailing pipes, blank lines between rows)
            with open(fpath, encoding='utf-8', errors='ignore') as f:
                raw_lines = f.readlines()

            # Strip leading/trailing pipes and split
            parsed = []
            for line in raw_lines:
                line = line.strip()
                if not line:
                    continue
                # Remove leading and trailing pipe
                if line.startswith('|'):
                    line = line[1:]
                if line.endswith('|'):
                    line = line[:-1]
                fields = [c.strip() for c in line.split('|')]
                parsed.append(fields)

            if len(parsed) < 2:
                raise ValueError("File has no data rows")

            header = parsed[0]
            # Find column indices
            col_map = {}
            for i, h in enumerate(header):
                hl = h.lower().replace(' ', '').replace('_', '')
                if hl in ('supplychain', 'supply_chain', 'supplychaingroup'):
                    col_map['supply_chain'] = i
                elif hl == 'fsc':
                    col_map['fsc'] = i
                elif hl == 'niin':
                    col_map['niin'] = i
                elif hl in ('itemdescription', 'item_description', 'nomenclature', 'desc'):
                    col_map['desc'] = i
                elif hl == 'ui':
                    col_map['ui'] = i
                elif hl == 'total':
                    col_map['total'] = i
            if 'niin' not in col_map:
                raise ValueError(f"Could not find NIIN column. Headers: {header}")

            # Identify month columns (MM/YYYY format)
            month_cols = []
            for i, h in enumerate(header):
                if re.match(r'^\d{2}/\d{4}$', h.strip()):
                    month_cols.append((i, h.strip()))

            for row in parsed[1:]:
                if len(row) < 3:
                    continue
                niin_raw = row[col_map['niin']] if col_map['niin'] < len(row) else ""
                niin = norm_niin(niin_raw)
                if not niin:
                    continue

                def _get(key):
                    idx = col_map.get(key)
                    return row[idx] if idx is not None and idx < len(row) else ""

                fsc = _get('fsc')
                sc = _get('supply_chain')
                desc = _get('desc')
                ui = _get('ui')

                # Parse monthly demands
                monthly = {}
                for ci, mn in month_cols:
                    if ci < len(row):
                        try:
                            v = int(float(row[ci])) if row[ci] else 0
                        except (ValueError, TypeError):
                            v = 0
                        monthly[mn] = v

                total = sum(monthly.values())
                # Override with Total column if present
                if col_map.get('total') is not None and col_map['total'] < len(row):
                    try:
                        total = int(float(row[col_map['total']]))
                    except (ValueError, TypeError):
                        pass

                nonzero_months = sum(1 for v in monthly.values() if v > 0)
                avg_monthly = total / max(1, len(monthly)) if monthly else 0
                peak = max(monthly.values()) if monthly else 0
                months_sorted = sorted(monthly.keys())

                rows_data.append((
                    niin, fsc, sc, desc, ui, total,
                    json.dumps(monthly), round(avg_monthly, 2), peak, nonzero_months,
                    months_sorted[0] if months_sorted else "",
                    months_sorted[-1] if months_sorted else "",
                    source_name
                ))

        elif fpath.suffix.lower() in ('.xls', '.xlsx', '.html', '.htm'):
            # HTML table disguised as .xls (DIBBS format)
            try:
                import pandas as pd
                dfs = pd.read_html(str(fpath))
                df = dfs[0]
                # Flatten multi-index if present
                if isinstance(df.columns, pd.MultiIndex):
                    df.columns = df.columns.get_level_values(-1)
                df.columns = [str(c).strip() for c in df.columns]
            except ImportError:
                # Fallback: try as pipe-delimited txt companion
                txt_path = fpath.with_suffix('.txt')
                if txt_path.exists():
                    db.close()
                    return run_forecast_import_worker(log_id, str(txt_path))
                raise ValueError("pandas not installed and no .txt companion file found. Install pandas: pip install pandas lxml")

            # Find columns
            niin_col = None
            fsc_col = None
            sc_col = None
            desc_col = None
            ui_col = None
            total_col = None
            month_cols = []

            for c in df.columns:
                cl = c.lower().replace(' ', '').replace('_', '')
                if cl == 'niin':
                    niin_col = c
                elif cl == 'fsc':
                    fsc_col = c
                elif cl in ('supplychain', 'supplychaingroup'):
                    sc_col = c
                elif cl in ('itemdescription', 'nomenclature', 'desc'):
                    desc_col = c
                elif cl == 'ui':
                    ui_col = c
                elif cl == 'total':
                    total_col = c
                elif re.match(r'^\d{2}/\d{4}$', c.strip()):
                    month_cols.append(c)

            if niin_col is None:
                raise ValueError(f"Could not find NIIN column. Columns: {list(df.columns)}")

            for _, row in df.iterrows():
                niin = norm_niin(row.get(niin_col))
                if not niin:
                    continue
                fsc = str(row.get(fsc_col, '')).strip() if fsc_col else ""
                sc = str(row.get(sc_col, '')).strip() if sc_col else ""
                desc = str(row.get(desc_col, '')).strip() if desc_col else ""
                ui = str(row.get(ui_col, '')).strip() if ui_col else ""

                monthly = {}
                for mc in month_cols:
                    try:
                        v = int(float(row.get(mc, 0)))
                    except (ValueError, TypeError):
                        v = 0
                    monthly[mc] = v

                total = sum(monthly.values())
                if total_col:
                    try:
                        total = int(float(row.get(total_col, total)))
                    except (ValueError, TypeError):
                        pass

                nonzero_months = sum(1 for v in monthly.values() if v > 0)
                avg_monthly = total / max(1, len(monthly)) if monthly else 0
                peak = max(monthly.values()) if monthly else 0
                months_sorted = sorted(monthly.keys())

                rows_data.append((
                    niin, fsc, sc, desc, ui, total,
                    json.dumps(monthly), round(avg_monthly, 2), peak, nonzero_months,
                    months_sorted[0] if months_sorted else "",
                    months_sorted[-1] if months_sorted else "",
                    source_name
                ))
        else:
            raise ValueError(f"Unsupported file format: {fpath.suffix}")

        # Write to DB
        db.execute("DELETE FROM forecast")
        for rd in rows_data:
            db.execute("""
                INSERT OR REPLACE INTO forecast
                (niin, fsc, supply_chain, item_description, ui, total_demand,
                 monthly_demand, avg_monthly, peak_monthly, demand_months,
                 forecast_start, forecast_end, source_file)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, rd)

        # Auto-populate niin_targets from forecast
        existing_targets = {r["niin"] for r in db.execute("SELECT niin FROM niin_targets").fetchall()}
        new_targets = 0
        for rd in rows_data:
            niin = rd[0]
            if niin not in existing_targets:
                db.execute("INSERT OR IGNORE INTO niin_targets (niin) VALUES (?)", (niin,))
                new_targets += 1

        db.commit()
        db.close()
        msg = f"Imported {len(rows_data)} NIINs from forecast. Added {new_targets} new target NIINs."
        _update_log(log_id, "success", len(rows_data), msg)
        log.info("Forecast import: %s", msg)
    except Exception as e:
        log.exception("Forecast import failed")
        _update_log(log_id, "error", log_text=str(e))


def build_niin_datasheet(db, niin: str) -> dict:
    """Build a comprehensive datasheet for a single NIIN from PUBLOG data."""
    ds = {"niin": niin}

    # Identification
    ident = db.execute("SELECT * FROM publog_identification WHERE niin=?", (niin,)).fetchone()
    if ident:
        ds["identification"] = dict(ident)
        # Resolve FSC title
        fsc_info = db.execute("SELECT * FROM publog_hseries_fsc WHERE fsc=?", (ident["fsc"],)).fetchone()
        ds["fsc_info"] = dict(fsc_info) if fsc_info else None
        # Resolve INC name/definition
        inc_info = db.execute("SELECT * FROM publog_hseries_inc WHERE inc=?", (ident["inc"],)).fetchone()
        ds["inc_info"] = dict(inc_info) if inc_info else None
    else:
        ds["identification"] = None
        ds["fsc_info"] = None
        ds["inc_info"] = None

    # Characteristics
    chars = db.execute("""
        SELECT mrc, requirements_statement, clear_text_reply
        FROM publog_characteristics WHERE niin=?
        ORDER BY mrc
    """, (niin,)).fetchall()
    ds["characteristics"] = [dict(c) for c in chars]

    # Management (most recent by effective_date)
    mgmt = db.execute("""
        SELECT * FROM publog_management WHERE niin=?
        ORDER BY effective_date DESC
    """, (niin,)).fetchall()
    ds["management"] = [dict(m) for m in mgmt]

    # References (part numbers)
    refs = db.execute("""
        SELECT r.*, c.company, c.city, c.state_province, c.country
        FROM publog_references r
        LEFT JOIN publog_cage c ON r.cage_code = c.cage_code
        WHERE r.niin=?
        ORDER BY r.cage_code, r.part_number
    """, (niin,)).fetchall()
    ds["references"] = [dict(r) for r in refs]

    # Summary stats
    mgmt_dicts = ds["management"]
    ds["stats"] = {
        "char_count": len(chars),
        "mgmt_records": len(mgmt_dicts),
        "part_numbers": len(ds["references"]),
        "unique_cages": len(set(r["cage_code"] for r in ds["references"] if r.get("cage_code"))),
        "has_unit_price": any(m.get("unit_price") for m in mgmt_dicts),
    }

    return ds


def seed_default_config():
    db = get_db()
    # Clean up old config keys no longer in DEFAULTS
    existing = {r["key"] for r in db.execute("SELECT key FROM config WHERE section='decision'").fetchall()}
    for old_key in existing - set(DEFAULTS.keys()):
        db.execute("DELETE FROM config WHERE key=? AND section='decision'", (old_key,))
    for k, v in DEFAULTS.items():
        db.execute("""
            INSERT OR IGNORE INTO config (key, value, section, description)
            VALUES (?,?,?,?)
        """, (k, str(v), "decision", DECISION_DESCRIPTIONS.get(k, "")))
    for k, v in FINGERPRINT_DEFAULTS.items():
        db.execute("""
            INSERT OR IGNORE INTO config (key, value, section, description)
            VALUES (?,?,?,?)
        """, (k, str(v), "fingerprint", FINGERPRINT_DESCRIPTIONS.get(k, "")))
    # Crawler config section
    for k, v in CRAWLER_DEFAULTS.items():
        db.execute("""
            INSERT OR IGNORE INTO config (key, value, section, description)
            VALUES (?,?,?,?)
        """, (k, str(v), "crawler", CRAWLER_DESCRIPTIONS.get(k, "")))
    # General config section (API keys etc.)
    db.execute("""
        INSERT OR IGNORE INTO config (key, value, section, description)
        VALUES ('GEMINI_API_KEY', 'AIzaSyCGe_seNc-BlyppbNuTxnd9K27TofAbn1c', 'general',
                'Google Gemini API key for AI image generation')
    """)
    db.commit()
    db.close()


# ──────────────────────────────────────────────
# SECTION H: Flask API Routes
# ──────────────────────────────────────────────

@app.errorhandler(500)
def handle_500(e):
    if request.path.startswith("/api/"):
        return jsonify({"error": str(e)}), 500
    return str(e), 500


@app.errorhandler(404)
def handle_404(e):
    if request.path.startswith("/api/"):
        return jsonify({"error": "not found"}), 404
    return render_template("index.html")


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/dashboard")
def api_dashboard():
    db = get_db()
    total = db.execute("SELECT COUNT(*) c FROM command_center").fetchone()["c"]
    actions = db.execute("""
        SELECT action, COUNT(*) c FROM command_center
        WHERE action != '' GROUP BY action ORDER BY c DESC
    """).fetchall()
    top10 = db.execute("""
        SELECT niin, action, opportunity_score, confidence
        FROM command_center WHERE opportunity_score > 0
        ORDER BY opportunity_score DESC LIMIT 10
    """).fetchall()

    avg_score = db.execute(
        "SELECT AVG(opportunity_score) a FROM command_center WHERE opportunity_score > 0"
    ).fetchone()["a"] or 0

    logs_today = db.execute("""
        SELECT COUNT(*) c FROM pipeline_log
        WHERE date(started_at) = date('now')
    """).fetchone()["c"]

    db.close()
    return jsonify({
        "total_niins": total,
        "avg_score": round(avg_score, 1),
        "pipelines_today": logs_today,
        "actions": [{"action": r["action"], "count": r["c"]} for r in actions],
        "top10": [dict(r) for r in top10],
    })


@app.route("/api/command-center")
def api_command_center():
    db = get_db()
    action_filter = request.args.get("action", "")
    search = request.args.get("search", "")
    sort = request.args.get("sort", "opportunity_score")
    order = request.args.get("order", "desc")
    page = max(1, int(request.args.get("page", 1)))
    per_page = min(200, max(10, int(request.args.get("per_page", 50))))

    allowed_sorts = {"niin", "opportunity_score", "action", "confidence", "coverage",
                     "fp_conf", "nobid_score", "median_unit_price", "supplier_count",
                     "evidence_score", "fp_score", "docs_score", "profit_score_val", "docs_total",
                     "demand_score", "char_score", "supply_score", "price_score", "competition_score",
                     "cots_readiness", "total_demand", "char_count", "ref_count", "cage_count",
                     "govt_unit_price", "item_name", "crit_cd", "dmil"}
    if sort not in allowed_sorts:
        sort = "opportunity_score"
    if order not in ("asc", "desc"):
        order = "desc"

    where_clauses = []
    params = []
    if action_filter:
        where_clauses.append("action = ?")
        params.append(action_filter)
    if search:
        where_clauses.append("niin LIKE ?")
        params.append(f"%{search}%")

    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

    total = db.execute(f"SELECT COUNT(*) c FROM command_center {where_sql}", params).fetchone()["c"]
    rows = db.execute(
        f"SELECT * FROM command_center {where_sql} ORDER BY {sort} {order} LIMIT ? OFFSET ?",
        params + [per_page, (page - 1) * per_page]
    ).fetchall()

    db.close()
    return jsonify({
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": math.ceil(total / per_page) if total else 1,
        "rows": [dict(r) for r in rows],
    })


@app.route("/api/niin/<niin>")
def api_niin_detail(niin):
    db = get_db()
    cc = db.execute("SELECT * FROM command_center WHERE niin=?", (niin,)).fetchone()
    pricing = db.execute("SELECT * FROM pricing_signals WHERE niin=?", (niin,)).fetchone()
    nobid = db.execute("SELECT * FROM nobid_signals WHERE niin=?", (niin,)).fetchone()
    fp = db.execute("SELECT * FROM fingerprint_by_niin WHERE niin=?", (niin,)).fetchone()
    docs = db.execute("SELECT * FROM doc_extract WHERE niin=? ORDER BY confidence DESC", (niin,)).fetchall()
    evidence = db.execute("SELECT * FROM evidence_links WHERE niin=?", (niin,)).fetchall()
    flis = db.execute("SELECT * FROM flis_refnums WHERE niin=?", (niin,)).fetchall()
    fc = db.execute("SELECT * FROM forecast WHERE niin=?", (niin,)).fetchone()
    db.close()
    return jsonify({
        "command_center": dict(cc) if cc else None,
        "pricing": dict(pricing) if pricing else None,
        "nobid": dict(nobid) if nobid else None,
        "fingerprint": dict(fp) if fp else None,
        "documents": [dict(d) for d in docs],
        "evidence_links": [dict(e) for e in evidence],
        "flis_refs": [dict(f) for f in flis],
        "forecast": dict(fc) if fc else None,
    })


@app.route("/api/config/<section>")
def api_get_config(section):
    db = get_db()
    rows = db.execute("SELECT key, value, description FROM config WHERE section=? ORDER BY key", (section,)).fetchall()
    db.close()
    return jsonify([{"key": r["key"], "value": r["value"], "description": r["description"]} for r in rows])


@app.route("/api/config/<section>", methods=["PUT"])
def api_put_config(section):
    items = request.get_json()
    db = get_db()
    for item in items:
        db.execute("""
            INSERT OR REPLACE INTO config (key, value, section, updated_at)
            VALUES (?, ?, ?, datetime('now'))
        """, (item["key"], item["value"], section))
    db.commit()
    db.close()
    return jsonify({"status": "ok", "updated": len(items)})


@app.route("/api/config/<section>/reset", methods=["POST"])
def api_reset_config(section):
    db = get_db()
    defaults_map = {"decision": DEFAULTS, "fingerprint": FINGERPRINT_DEFAULTS}
    desc_map = {"decision": DECISION_DESCRIPTIONS, "fingerprint": FINGERPRINT_DESCRIPTIONS}
    defaults = defaults_map.get(section, {})
    descs = desc_map.get(section, {})
    db.execute("DELETE FROM config WHERE section=?", (section,))
    for k, v in defaults.items():
        db.execute("""
            INSERT INTO config (key, value, section, description) VALUES (?,?,?,?)
        """, (k, str(v), section, descs.get(k, "")))
    db.commit()
    db.close()
    return jsonify({"status": "ok", "reset": len(defaults)})


@app.route("/api/logs")
def api_logs():
    db = get_db()
    rows = db.execute("SELECT * FROM pipeline_log ORDER BY id DESC LIMIT 100").fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/pipeline-status/<int:log_id>")
def api_pipeline_status(log_id):
    db = get_db()
    row = db.execute("SELECT * FROM pipeline_log WHERE id=?", (log_id,)).fetchone()
    db.close()
    if not row:
        return jsonify({"error": "not found"}), 404
    return jsonify(dict(row))


# ── Import endpoints ──

def _start_pipeline(name, worker_fn, *args):
    db = get_db()
    log_id = db.execute(
        "INSERT INTO pipeline_log (pipeline_name, status) VALUES (?, 'running')", (name,)
    ).lastrowid
    db.commit()
    db.close()
    threading.Thread(target=worker_fn, args=(log_id, *args), daemon=True).start()
    return log_id


@app.route("/api/import/pricing", methods=["POST"])
def api_import_pricing():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    filename = secure_filename(f.filename)
    save_path = str(UPLOAD_DIR / f"pricing_{int(time.time())}_{filename}")
    f.save(save_path)
    sheet = request.form.get("sheet", "FY26 Pricing")
    log_id = _start_pipeline("import_pricing", run_import_pricing_worker, save_path, sheet)
    return jsonify({"job_id": log_id, "status": "running"})


@app.route("/api/import/flis", methods=["POST"])
def api_import_flis():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    filename = secure_filename(f.filename)
    save_path = str(UPLOAD_DIR / f"flis_{int(time.time())}_{filename}")
    f.save(save_path)
    file_type = "csv" if filename.lower().endswith(".csv") else "xlsx"
    sheet = request.form.get("sheet", "Sheet1")
    log_id = _start_pipeline("import_flis", run_flis_import_worker, save_path, file_type, sheet)
    return jsonify({"job_id": log_id, "status": "running"})


@app.route("/api/import/nobid", methods=["POST"])
def api_import_nobid():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    filename = secure_filename(f.filename)
    save_path = str(UPLOAD_DIR / f"nobid_{int(time.time())}_{filename}")
    f.save(save_path)
    log_id = _start_pipeline("import_nobid", run_import_nobid_worker, save_path)
    return jsonify({"job_id": log_id, "status": "running"})


@app.route("/api/import/pdfs", methods=["POST"])
def api_import_pdfs():
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "No files uploaded"}), 400
    saved = 0
    for f in files:
        if f.filename and f.filename.lower().endswith(".pdf"):
            fname = secure_filename(f.filename)
            f.save(str(PDF_DIR / fname))
            saved += 1
    return jsonify({"status": "ok", "saved": saved})


@app.route("/api/import/excel-migration", methods=["POST"])
def api_excel_migration():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    filename = secure_filename(f.filename)
    save_path = str(UPLOAD_DIR / f"migration_{int(time.time())}_{filename}")
    f.save(save_path)
    log_id = _start_pipeline("excel_migration", run_excel_migration_worker, save_path)
    return jsonify({"job_id": log_id, "status": "running"})


# ── Run pipeline endpoints ──

@app.route("/api/run/decision-engine", methods=["POST"])
def api_run_decision_engine():
    log_id = _start_pipeline("decision_engine", run_decision_engine_worker)
    return jsonify({"job_id": log_id, "status": "running"})


@app.route("/api/run/fingerprint", methods=["POST"])
def api_run_fingerprint():
    log_id = _start_pipeline("fingerprint", run_fingerprint_worker)
    return jsonify({"job_id": log_id, "status": "running"})


# ── PUBLOG endpoints ──

PUBLOG_DIR = BASE_DIR / "PUBLOG"

@app.route("/api/import/publog", methods=["POST"])
def api_import_publog():
    data = request.json if request.is_json else {}
    publog_path = data.get("path", str(PUBLOG_DIR))
    min_demand = int(data.get("min_demand", 0))
    if not Path(publog_path).exists():
        return jsonify({"error": f"PUBLOG directory not found: {publog_path}"}), 400
    log_id = _start_pipeline("publog_import", run_publog_import_worker, publog_path, min_demand)
    return jsonify({"job_id": log_id, "status": "running"})


@app.route("/api/publog/status")
def api_publog_status():
    db = get_db()
    counts = {}
    for tbl in ("publog_identification", "publog_characteristics", "publog_management",
                "publog_references", "publog_cage", "publog_hseries_fsc", "publog_hseries_inc"):
        counts[tbl] = db.execute(f"SELECT COUNT(*) c FROM {tbl}").fetchone()["c"]
    db.close()
    return jsonify(counts)


@app.route("/api/niin/<niin>/datasheet")
def api_niin_datasheet(niin):
    db = get_db()
    ds = build_niin_datasheet(db, niin)
    db.close()
    return jsonify(ds)


@app.route("/api/import/forecast", methods=["POST"])
def api_import_forecast():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    filename = secure_filename(f.filename)
    save_path = str(UPLOAD_DIR / f"forecast_{int(time.time())}_{filename}")
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    f.save(save_path)
    log_id = _start_pipeline("forecast_import", run_forecast_import_worker, save_path)
    return jsonify({"job_id": log_id, "status": "running"})


@app.route("/api/forecast")
def api_forecast():
    """Return forecast data with optional filters and sorting."""
    db = get_db()
    sort = request.args.get("sort", "total_demand")
    order = request.args.get("order", "desc")
    search = request.args.get("search", "").strip()
    supply_chain = request.args.get("supply_chain", "")
    min_demand = request.args.get("min_demand", "0")
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 50))

    allowed_sorts = {"niin", "fsc", "total_demand", "avg_monthly", "peak_monthly",
                     "demand_months", "item_description", "supply_chain"}
    if sort not in allowed_sorts:
        sort = "total_demand"
    order = "DESC" if order.lower() == "desc" else "ASC"

    where = ["1=1"]
    params = []
    if search:
        where.append("(f.niin LIKE ? OR f.item_description LIKE ?)")
        params.extend([f"%{search}%", f"%{search}%"])
    if supply_chain:
        where.append("f.supply_chain = ?")
        params.append(supply_chain)
    try:
        md = int(min_demand)
        if md > 0:
            where.append("f.total_demand >= ?")
            params.append(md)
    except (ValueError, TypeError):
        pass

    where_sql = " AND ".join(where)
    total = db.execute(f"SELECT COUNT(*) c FROM forecast f WHERE {where_sql}", params).fetchone()["c"]
    pages = max(1, (total + per_page - 1) // per_page)

    rows = db.execute(f"""
        SELECT f.*,
            p.median_unit_price, p.last_unit_price, p.supplier_count, p.total_awards,
            cc.action, cc.opportunity_score,
            pm.unit_price as govt_unit_price
        FROM forecast f
        LEFT JOIN pricing_signals p ON f.niin = p.niin
        LEFT JOIN command_center cc ON f.niin = cc.niin
        LEFT JOIN (
            SELECT niin, unit_price FROM publog_management
            WHERE id IN (SELECT MIN(id) FROM publog_management GROUP BY niin)
        ) pm ON f.niin = pm.niin
        WHERE {where_sql}
        ORDER BY f.{sort} {order}
        LIMIT ? OFFSET ?
    """, params + [per_page, (page - 1) * per_page]).fetchall()

    # Supply chains for filter dropdown
    chains = db.execute("SELECT DISTINCT supply_chain FROM forecast WHERE supply_chain != '' ORDER BY supply_chain").fetchall()

    db.close()
    return jsonify({
        "rows": [dict(r) for r in rows],
        "total": total, "page": page, "pages": pages,
        "supply_chains": [c["supply_chain"] for c in chains],
    })


@app.route("/api/forecast/summary")
def api_forecast_summary():
    db = get_db()
    total_niins = db.execute("SELECT COUNT(*) c FROM forecast").fetchone()["c"]
    total_demand = db.execute("SELECT COALESCE(SUM(total_demand),0) s FROM forecast").fetchone()["s"]
    avg_demand = db.execute("SELECT COALESCE(AVG(total_demand),0) a FROM forecast").fetchone()["a"]
    with_pricing = db.execute("""
        SELECT COUNT(*) c FROM forecast f
        INNER JOIN pricing_signals p ON f.niin = p.niin
    """).fetchone()["c"]
    with_publog = db.execute("""
        SELECT COUNT(*) c FROM forecast f
        INNER JOIN publog_identification pi ON f.niin = pi.niin
    """).fetchone()["c"]
    by_chain = db.execute("""
        SELECT supply_chain, COUNT(*) cnt, SUM(total_demand) demand
        FROM forecast GROUP BY supply_chain ORDER BY demand DESC
    """).fetchall()
    high_demand = db.execute("""
        SELECT f.niin, f.item_description, f.total_demand, f.supply_chain, f.ui,
            p.median_unit_price, cc.action
        FROM forecast f
        LEFT JOIN pricing_signals p ON f.niin = p.niin
        LEFT JOIN command_center cc ON f.niin = cc.niin
        ORDER BY f.total_demand DESC LIMIT 20
    """).fetchall()
    db.close()
    return jsonify({
        "total_niins": total_niins,
        "total_demand": total_demand,
        "avg_demand": round(avg_demand, 1),
        "with_pricing": with_pricing,
        "with_publog": with_publog,
        "by_chain": [dict(r) for r in by_chain],
        "high_demand": [dict(r) for r in high_demand],
    })


@app.route("/api/datasheet/bulk")
def api_datasheet_bulk():
    """Return summary datasheet info for NIINs that have PUBLOG data, with filtering and pagination."""
    db = get_db()
    search = request.args.get("search", "").strip()
    min_demand = request.args.get("min_demand", "0")
    sort = request.args.get("sort", "niin")
    order = request.args.get("order", "asc")
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 50))

    allowed_sorts = {"niin", "item_name", "fsc", "govt_price", "char_count", "ref_count", "total_demand"}
    if sort not in allowed_sorts:
        sort = "niin"
    order = "DESC" if order.lower() == "desc" else "ASC"

    where = ["1=1"]
    params = []
    if search:
        where.append("(p.niin LIKE ? OR p.item_name LIKE ?)")
        params.extend([f"%{search}%", f"%{search}%"])
    try:
        md = int(min_demand)
        if md > 0:
            where.append("COALESCE(fc.total_demand, 0) >= ?")
            params.append(md)
    except (ValueError, TypeError):
        pass

    where_sql = " AND ".join(where)
    total = db.execute(f"""
        SELECT COUNT(*) c FROM publog_identification p
        LEFT JOIN forecast fc ON p.niin = fc.niin
        WHERE {where_sql}
    """, params).fetchone()["c"]
    pages = max(1, (total + per_page - 1) // per_page)

    rows = db.execute(f"""
        SELECT p.niin, p.fsc, p.inc, p.item_name, p.sos, p.dmil,
               f.fsc_title,
               (SELECT COUNT(*) FROM publog_characteristics c WHERE c.niin=p.niin) as char_count,
               (SELECT COUNT(*) FROM publog_references r WHERE r.niin=p.niin) as ref_count,
               (SELECT unit_price FROM publog_management m WHERE m.niin=p.niin ORDER BY effective_date DESC LIMIT 1) as govt_price,
               (SELECT ui FROM publog_management m WHERE m.niin=p.niin ORDER BY effective_date DESC LIMIT 1) as ui,
               fc.total_demand, fc.supply_chain,
               ps.median_unit_price,
               cc.action, cc.opportunity_score
        FROM publog_identification p
        LEFT JOIN publog_hseries_fsc f ON p.fsc = f.fsc
        LEFT JOIN forecast fc ON p.niin = fc.niin
        LEFT JOIN pricing_signals ps ON p.niin = ps.niin
        LEFT JOIN command_center cc ON p.niin = cc.niin
        WHERE {where_sql}
        ORDER BY {('fc.' if sort == 'total_demand' else 'p.' if sort in ('niin','item_name','fsc') else '')}{sort} {order}
        LIMIT ? OFFSET ?
    """, params + [per_page, (page - 1) * per_page]).fetchall()
    db.close()
    return jsonify({
        "rows": [dict(r) for r in rows],
        "total": total, "page": page, "pages": pages,
    })


# ──────────────────────────────────────────────
# SECTION I-A: COTS Assessment API
# ──────────────────────────────────────────────

@app.route("/api/cots-assessment/<niin>")
def api_cots_assessment(niin):
    """Full COTS assessment for a single NIIN."""
    db = get_db()
    cfg = load_config_from_db(db, "decision", DEFAULTS)

    # Characteristics
    chars = db.execute("""
        SELECT mrc, requirements_statement, clear_text_reply
        FROM publog_characteristics WHERE niin=? ORDER BY mrc
    """, (niin,)).fetchall()
    chars_list = [dict(c) for c in chars]
    char_count = len(chars_list)

    # References with CAGE company info
    refs = db.execute("""
        SELECT r.part_number, r.cage_code, r.cage_status, r.rncc, r.rnvc, r.dac,
               c.company, c.city, c.state_province, c.country, c.type as cage_type
        FROM publog_references r
        LEFT JOIN publog_cage c ON r.cage_code = c.cage_code
        WHERE r.niin=? ORDER BY r.cage_code
    """, (niin,)).fetchall()
    refs_list = [dict(r) for r in refs]
    unique_cages = list({r["cage_code"] for r in refs_list if r.get("cage_code")})
    ref_count = len(refs_list)
    cage_count = len(unique_cages)

    # Group parts by supplier
    suppliers = {}
    for r in refs_list:
        cage = r.get("cage_code", "")
        if cage not in suppliers:
            suppliers[cage] = {"cage": cage, "company": r.get("company", ""),
                               "city": r.get("city", ""), "state": r.get("state_province", ""),
                               "country": r.get("country", ""), "parts": []}
        suppliers[cage]["parts"].append(r.get("part_number", ""))

    # Identification
    ident = db.execute("SELECT * FROM publog_identification WHERE niin=?", (niin,)).fetchone()
    crit_cd = ident["crit_cd"] if ident else ""
    dmil_val = ident["dmil"] if ident else ""
    item_name = ident["item_name"] if ident else ""

    # Forecast
    fc = db.execute("SELECT total_demand, avg_monthly FROM forecast WHERE niin=?", (niin,)).fetchone()
    total_demand = fc["total_demand"] if fc else 0

    # Compute scores
    char_s = calc_char_score(char_count)
    supply_s = calc_supply_score(cage_count, ref_count)
    cots_r = calc_cots_readiness(char_s, supply_s, dmil_val, crit_cd, cfg)

    # Restrictions
    restrictions = {"dmil": dmil_val, "crit_cd": crit_cd, "flags": []}
    if dmil_val and dmil_val.strip():
        restrictions["flags"].append(f"DMIL code {dmil_val} — may restrict COTS sourcing")
    if crit_cd in ('C', 'D'):
        restrictions["flags"].append(f"Criticality code {crit_cd} — requires additional qualification")

    # SAR Rating
    sar_factors = []
    if char_count >= 10:
        sar_factors.append("Well-defined item (10+ characteristics)")
    elif char_count >= 5:
        sar_factors.append("Moderately defined item (5-9 characteristics)")
    else:
        sar_factors.append(f"Poorly defined item (only {char_count} characteristics)")

    if cage_count >= 2:
        sar_factors.append(f"Multiple known sources ({cage_count} CAGE codes)")
    elif cage_count == 1:
        sar_factors.append("Single known source")
    else:
        sar_factors.append("No known OEM sources in PUBLOG")

    if dmil_val and dmil_val.strip():
        sar_factors.append(f"DMIL restriction: {dmil_val}")
    else:
        sar_factors.append("No DMIL restrictions")

    if total_demand > 50:
        sar_factors.append(f"Strong demand ({total_demand} units)")
    elif total_demand > 0:
        sar_factors.append(f"Moderate demand ({total_demand} units)")

    if char_count >= 10 and cage_count >= 2 and (not dmil_val or not dmil_val.strip()) and total_demand > 50:
        sar_rating = "HIGH"
    elif char_count >= 5 and cage_count >= 1:
        sar_rating = "MEDIUM"
    else:
        sar_rating = "LOW"

    # Drop-in Replacement Assessment
    dropin_factors = []
    if ref_count >= 3:
        dropin_factors.append(f"Multiple known part numbers ({ref_count}) — good cross-reference potential")
    elif ref_count > 0:
        dropin_factors.append(f"Limited part numbers ({ref_count})")
    else:
        dropin_factors.append("No known part numbers — direct replacement identification difficult")

    if char_count >= 15:
        dropin_factors.append("Highly specified item — precise matching required but well-documented")
    elif char_count >= 5:
        dropin_factors.append("Moderately specified — reasonable matching possible")

    # Check for dimensional/standard specs in characteristics
    has_dimensional = False
    has_standard = False
    for c in chars_list:
        stmt = (c.get("requirements_statement") or "").lower()
        reply = (c.get("clear_text_reply") or "").lower()
        combined = stmt + " " + reply
        if any(k in combined for k in ("thread", "diameter", "length", "width", "height", "bore", "od", "id")):
            has_dimensional = True
        if any(k in combined for k in ("mil-", "sae", "astm", "an ", "nas ", "ams")):
            has_standard = True

    if has_dimensional:
        dropin_factors.append("Has dimensional specifications — supports physical matching")
    if has_standard:
        dropin_factors.append("References industry/military standards — supports spec-based matching")

    if ref_count >= 3 and has_dimensional and cage_count >= 2:
        dropin_rating = "HIGH"
    elif ref_count >= 1 and (has_dimensional or char_count >= 5):
        dropin_rating = "MEDIUM"
    else:
        dropin_rating = "LOW"

    db.close()
    return jsonify({
        "cots_readiness": round(cots_r, 1),
        "char_count": char_count,
        "ref_count": ref_count,
        "cage_count": cage_count,
        "item_name": item_name,
        "characteristics": chars_list,
        "suppliers": list(suppliers.values()),
        "restrictions": restrictions,
        "sar_rating": sar_rating,
        "sar_factors": sar_factors,
        "dropin_rating": dropin_rating,
        "dropin_factors": dropin_factors,
    })


# ──────────────────────────────────────────────
# SECTION I-A-2: COTS Profit Assessment
# ──────────────────────────────────────────────

def _search_cots_alternatives_gemini(item_name, part_numbers, characteristics):
    """Use Gemini to find commercial alternatives for a military part."""
    try:
        from google import genai
        db = get_db()
        api_key_row = db.execute(
            "SELECT value FROM config WHERE key='GEMINI_API_KEY' AND section='general'"
        ).fetchone()
        db.close()
        if not api_key_row or not api_key_row["value"]:
            return []

        client = genai.Client(api_key=api_key_row["value"])

        # Build a description from characteristics
        char_lines = []
        for c in characteristics[:15]:
            stmt = c.get("requirements_statement", "")
            reply = c.get("clear_text_reply", "")
            if stmt and reply:
                char_lines.append(f"- {stmt}: {reply}")
        char_text = "\n".join(char_lines) if char_lines else "No detailed specifications available."

        pn_text = ", ".join(part_numbers[:5]) if part_numbers else "None"

        prompt = f"""You are a defense supply chain analyst. Find commercial off-the-shelf (COTS) alternatives for this military part:

Item: {item_name}
Known Part Numbers: {pn_text}
Specifications:
{char_text}

Find 3-5 commercially available alternatives. For each, provide:
1. Product name
2. Manufacturer/brand
3. Estimated price (USD) - give your best estimate based on similar products
4. Key specifications that match
5. A URL where it could be purchased (use real product URLs from McMaster-Carr, Grainger, MSC Industrial, etc.)

IMPORTANT: Respond ONLY with a valid JSON array. No markdown, no code blocks, no explanation.
Format: [{{"name":"...","manufacturer":"...","estimated_price":99.99,"specs":"...","url":"...","match_quality":"high/medium/low"}}]
If you cannot find alternatives, return an empty array: []"""

        response = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=prompt,
        )

        text = response.text.strip()
        # Clean up common Gemini formatting
        if text.startswith("```"):
            text = text.split("\n", 1)[1] if "\n" in text else text[3:]
            if text.endswith("```"):
                text = text[:-3]
            text = text.strip()
            if text.startswith("json"):
                text = text[4:].strip()

        import json as _json
        alternatives = _json.loads(text)
        if not isinstance(alternatives, list):
            return []
        # Validate and clean each alternative
        cleaned = []
        for alt in alternatives[:5]:
            if not isinstance(alt, dict):
                continue
            cleaned.append({
                "name": str(alt.get("name", "Unknown"))[:200],
                "manufacturer": str(alt.get("manufacturer", ""))[:100],
                "estimated_price": float(alt.get("estimated_price", 0)) if alt.get("estimated_price") else None,
                "specs": str(alt.get("specs", ""))[:500],
                "url": str(alt.get("url", ""))[:500],
                "match_quality": str(alt.get("match_quality", "medium"))[:10],
            })
        return cleaned
    except Exception as e:
        log.warning("COTS alternative search failed: %s", str(e)[:200])
        return []


def run_profit_assessment_worker(log_id, niin):
    """Generate profit assessment for a NIIN."""
    try:
        log.info("Starting profit assessment for NIIN %s", niin)
        db = _worker_db()

        # Get identification
        ident = db.execute("SELECT * FROM publog_identification WHERE niin=?", (niin,)).fetchone()
        item_name = ident["item_name"] if ident else "UNKNOWN"

        # Get characteristics
        chars = db.execute("""
            SELECT mrc, requirements_statement, clear_text_reply
            FROM publog_characteristics WHERE niin=? ORDER BY mrc
        """, (niin,)).fetchall()
        chars_list = [dict(c) for c in chars]

        # Get part numbers
        refs = db.execute("SELECT DISTINCT part_number FROM publog_references WHERE niin=?", (niin,)).fetchall()
        pns = [r["part_number"] for r in refs if r["part_number"]]

        # Get government pricing
        pricing = db.execute("SELECT * FROM pricing_signals WHERE niin=?", (niin,)).fetchone()
        govt_price = None
        if pricing:
            govt_price = pricing["median_unit_price"] or pricing["last_unit_price"]

        # If no pricing_signals, check command_center
        if govt_price is None:
            cc = db.execute("SELECT govt_unit_price, median_unit_price FROM command_center WHERE niin=?", (niin,)).fetchone()
            if cc:
                govt_price = cc["govt_unit_price"] or cc["median_unit_price"]

        # Get forecast demand
        fc = db.execute("SELECT total_demand, avg_monthly FROM forecast WHERE niin=?", (niin,)).fetchone()
        total_demand = fc["total_demand"] if fc else 0
        avg_monthly = fc["avg_monthly"] if fc else 0
        annual_demand = int(avg_monthly * 12) if avg_monthly else total_demand

        db.close()

        # Search for COTS alternatives using Gemini
        log.info("Searching COTS alternatives for NIIN %s (%s)", niin, item_name)
        alternatives = _search_cots_alternatives_gemini(item_name, pns, chars_list)
        log.info("Found %d COTS alternatives for NIIN %s", len(alternatives), niin)

        # Calculate profit potential
        best_cots_price = None
        best_cots_source = ""
        for alt in alternatives:
            if alt.get("estimated_price") and alt["estimated_price"] > 0:
                if best_cots_price is None or alt["estimated_price"] < best_cots_price:
                    best_cots_price = alt["estimated_price"]
                    best_cots_source = alt.get("manufacturer", "") or alt.get("name", "")

        potential_margin = None
        roi_percent = None
        annual_revenue = None
        notes = []

        if govt_price and govt_price > 0 and best_cots_price and best_cots_price > 0:
            potential_margin = govt_price - best_cots_price
            if best_cots_price > 0:
                roi_percent = ((govt_price - best_cots_price) / best_cots_price) * 100
            annual_revenue = potential_margin * annual_demand if annual_demand > 0 else None

            if potential_margin > 0:
                notes.append(f"Potential margin of ${potential_margin:.2f} per unit")
            else:
                notes.append("COTS pricing may exceed government pricing — verify quotes")

            if roi_percent and roi_percent > 100:
                notes.append("High ROI potential — strong profit opportunity")
            elif roi_percent and roi_percent > 30:
                notes.append("Moderate ROI — worth pursuing with volume")

            if annual_demand > 100:
                notes.append(f"Strong demand ({annual_demand} units/year) supports investment")
        else:
            if not govt_price:
                notes.append("No government pricing data — import pricing file for full analysis")
            if not best_cots_price:
                notes.append("No COTS pricing found — manual research recommended")

        if not alternatives:
            notes.append("No commercial alternatives identified by AI — try manual search")

        # Save to DB
        db = _worker_db()
        db.execute("""
            INSERT OR REPLACE INTO profit_assessments
            (niin, item_name, govt_unit_price, cots_alternatives, best_cots_price,
             best_cots_source, potential_margin, roi_percent, annual_demand,
             annual_revenue_potential, assessment_notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (niin, item_name, govt_price, json.dumps(alternatives),
              best_cots_price, best_cots_source, potential_margin,
              roi_percent, annual_demand, annual_revenue,
              " | ".join(notes)))
        db.commit()
        db.close()

        _update_log(log_id, "success", 1,
                    f"Assessed NIIN {niin}: {len(alternatives)} alternatives found" +
                    (f", margin=${potential_margin:.2f}" if potential_margin else ""))

    except Exception as e:
        log.exception("Profit assessment failed for NIIN %s", niin)
        _update_log(log_id, "error", log_text=str(e)[:500])


@app.route("/api/niin/<niin>/profit-assessment", methods=["POST"])
def api_generate_profit_assessment(niin):
    """Start profit assessment for a NIIN."""
    log_id = _start_pipeline("profit_assessment", run_profit_assessment_worker, niin)
    return jsonify({"job_id": log_id, "status": "running"})


@app.route("/api/niin/<niin>/profit-assessment")
def api_get_profit_assessment(niin):
    """Get existing profit assessment for a NIIN."""
    db = get_db()
    row = db.execute("SELECT * FROM profit_assessments WHERE niin=?", (niin,)).fetchone()
    db.close()
    if not row:
        return jsonify({"exists": False})
    d = dict(row)
    d["cots_alternatives"] = json.loads(d.get("cots_alternatives", "[]"))
    d["exists"] = True
    return jsonify(d)


# ──────────────────────────────────────────────
# SECTION I-B: Crawler Feature
# ──────────────────────────────────────────────

def run_crawler_import_sites_worker(log_id, file_path):
    """Import crawler starter sites from Excel."""
    try:
        import openpyxl
        db = _worker_db()
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Empty file")

        header = [str(h).strip().lower() if h else "" for h in rows[0]]
        col_map = {}
        for i, h in enumerate(header):
            if "rank" in h: col_map["rank"] = i
            elif "site" in h and "name" in h: col_map["site_name"] = i
            elif "url" in h: col_map["url"] = i
            elif "value" in h or "primary" in h: col_map["primary_value"] = i
            elif "roi" in h or "expected" in h: col_map["expected_roi"] = i

        db.execute("DELETE FROM crawler_sites")
        count = 0
        for row in rows[1:]:
            if not row or all(c is None for c in row):
                continue
            rank = row[col_map.get("rank", 0)] if col_map.get("rank") is not None else count + 1
            site_name = str(row[col_map.get("site_name", 1)] or "").strip()
            url = str(row[col_map.get("url", 2)] or "").strip()
            pv = str(row[col_map.get("primary_value", 3)] or "").strip()
            roi = str(row[col_map.get("expected_roi", 4)] or "").strip()
            if not url or url.lower() == "various":
                continue
            try:
                rank_int = int(rank) if rank else count + 1
            except (ValueError, TypeError):
                rank_int = count + 1
            db.execute("""
                INSERT INTO crawler_sites (rank, site_name, url, primary_value, expected_roi)
                VALUES (?,?,?,?,?)
            """, (rank_int, site_name, url, pv, roi))
            count += 1

        db.commit()
        db.close()
        _update_log(log_id, "success", count, f"Imported {count} crawler sites.")
    except Exception as e:
        log.exception("Crawler sites import failed")
        _update_log(log_id, "error", log_text=str(e))


def run_crawler_worker(log_id, max_niins=50, max_pages_per_site=3):
    """Crawl enabled sites for target NIINs/PNs, download files, and store results."""
    try:
        db = _worker_db()

        sites = db.execute("SELECT * FROM crawler_sites WHERE enabled=1 ORDER BY rank").fetchall()
        if not sites:
            _update_log(log_id, "success", 0, "No crawler sites configured. Import sites first.")
            db.close()
            return

        niins_with_refs = db.execute("""
            SELECT f.niin, f.item_description,
                   GROUP_CONCAT(DISTINCT pr.part_number) as part_numbers
            FROM forecast f
            LEFT JOIN publog_references pr ON f.niin = pr.niin
            GROUP BY f.niin
            ORDER BY f.total_demand DESC
            LIMIT ?
        """, (max_niins,)).fetchall()

        if not niins_with_refs:
            _update_log(log_id, "success", 0, "No forecast NIINs to crawl.")
            db.close()
            return

        session = _requests_lib.Session()
        session.headers.update({"User-Agent": "Mozilla/5.0 (compatible; NSNIntelBot/1.0; research)",
                                "Accept": "text/html,application/xhtml+xml"})
        dl_cfg = load_config_from_db(db, "crawler", CRAWLER_DEFAULTS)

        total_results = 0
        total_files = 0
        errors = []

        for niin_row in niins_with_refs:
            niin = niin_row["niin"]
            desc = niin_row["item_description"] or ""
            pns = [p.strip() for p in (niin_row["part_numbers"] or "").split(",") if p.strip()]
            try:
                r_cnt, f_cnt = _crawl_niin_on_sites(db, session, niin, desc, pns, sites, {}, dl_cfg)
                total_results += r_cnt
                total_files += f_cnt
                db.commit()
            except Exception as ex:
                errors.append(f"{niin}: {str(ex)[:100]}")

        db.close()
        err_summary = f" Errors: {len(errors)}" if errors else ""
        _update_log(log_id, "success", total_results,
                    f"Crawled {len(sites)} sites for {len(niins_with_refs)} NIINs. "
                    f"{total_results} results, {total_files} files downloaded.{err_summary}")
    except Exception as e:
        log.exception("Crawler failed")
        _update_log(log_id, "error", log_text=str(e))


@app.route("/api/import/crawler-sites", methods=["POST"])
def api_import_crawler_sites():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    filename = secure_filename(f.filename)
    save_path = str(UPLOAD_DIR / f"crawler_{int(time.time())}_{filename}")
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    f.save(save_path)
    log_id = _start_pipeline("crawler_sites_import", run_crawler_import_sites_worker, save_path)
    return jsonify({"job_id": log_id, "status": "running"})


@app.route("/api/crawler/sites")
def api_crawler_sites():
    db = get_db()
    rows = db.execute("SELECT * FROM crawler_sites ORDER BY rank").fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/crawler/sites/<int:site_id>/toggle", methods=["POST"])
def api_toggle_crawler_site(site_id):
    db = get_db()
    db.execute("UPDATE crawler_sites SET enabled = CASE WHEN enabled=1 THEN 0 ELSE 1 END WHERE id=?", (site_id,))
    db.commit()
    db.close()
    return jsonify({"status": "ok"})


@app.route("/api/run/crawler", methods=["POST"])
def api_run_crawler():
    data = request.json if request.is_json else {}
    max_niins = int(data.get("max_niins", 50))
    max_pages = int(data.get("max_pages_per_site", 3))
    log_id = _start_pipeline("crawler", run_crawler_worker, max_niins, max_pages)
    return jsonify({"job_id": log_id, "status": "running"})


@app.route("/api/crawler/results")
def api_crawler_results():
    db = get_db()
    niin = request.args.get("niin", "")
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 50))
    where = ["1=1"]
    params = []
    if niin:
        where.append("niin = ?")
        params.append(niin)
    where_sql = " AND ".join(where)
    total = db.execute(f"SELECT COUNT(*) c FROM crawler_results WHERE {where_sql}", params).fetchone()["c"]
    rows = db.execute(f"""
        SELECT * FROM crawler_results WHERE {where_sql}
        ORDER BY relevance_score DESC
        LIMIT ? OFFSET ?
    """, params + [per_page, (page - 1) * per_page]).fetchall()
    db.close()
    return jsonify({
        "rows": [dict(r) for r in rows],
        "total": total, "page": page,
        "pages": max(1, (total + per_page - 1) // per_page),
    })


# ──────────────────────────────────────────────
# SECTION I-B-2: Crawler File Downloads & Document Intelligence
# ──────────────────────────────────────────────

def _detect_downloadable_links(soup, base_url):
    """Scan page for downloadable file links (PDFs, images, docs)."""
    import urllib.parse
    EXTENSIONS = {
        'pdf': ('pdf',),
        'image': ('jpg', 'jpeg', 'png', 'gif', 'tiff', 'bmp', 'webp'),
        'document': ('doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx'),
    }
    links = []
    seen = set()
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not href or href.startswith("#") or href.startswith("javascript"):
            continue
        full_url = urllib.parse.urljoin(base_url, href)
        if full_url in seen:
            continue
        ext = full_url.rsplit(".", 1)[-1].lower().split("?")[0] if "." in full_url else ""
        file_type = None
        for ftype, exts in EXTENSIONS.items():
            if ext in exts:
                file_type = ftype
                break
        if file_type:
            fname = urllib.parse.unquote(full_url.rsplit("/", 1)[-1].split("?")[0])
            is_datasheet = "datasheet" in fname.lower() or "spec" in fname.lower()
            links.append({"url": full_url, "file_type": file_type, "filename": fname[:200],
                          "is_datasheet": is_datasheet})
            seen.add(full_url)
    return links


def _download_crawler_file(session, url, niin, max_size_bytes, dl_dir):
    """Download a single file. Returns dict or None on failure."""
    fpath = None
    try:
        niin_dir = dl_dir / niin
        niin_dir.mkdir(parents=True, exist_ok=True)
        resp = session.get(url, stream=True, timeout=30)
        if resp.status_code != 200:
            log.debug("Download %s returned status %d", url[:120], resp.status_code)
            return None
        # Check Content-Length before downloading
        cl = resp.headers.get("Content-Length")
        if cl and int(cl) > max_size_bytes:
            log.debug("Download %s too large (%s bytes)", url[:120], cl)
            return None
        # Extract filename
        import urllib.parse
        raw_name = urllib.parse.unquote(url.rsplit("/", 1)[-1].split("?")[0])
        fname = secure_filename(raw_name) or f"file_{int(time.time())}"
        fname = f"{int(time.time())}_{fname}"
        fpath = niin_dir / fname
        total_size = 0
        with open(str(fpath), "wb") as f:
            for chunk in resp.iter_content(chunk_size=65536):
                total_size += len(chunk)
                if total_size > max_size_bytes:
                    break
                f.write(chunk)
        if total_size > max_size_bytes:
            fpath.unlink(missing_ok=True)
            return None
        if total_size == 0:
            fpath.unlink(missing_ok=True)
            return None
        mime = resp.headers.get("Content-Type", "").split(";")[0].strip()
        log.info("Downloaded %s (%d bytes) for NIIN %s", fname, total_size, niin)
        return {"filename": fname, "file_path": str(fpath), "file_size": total_size, "mime_type": mime}
    except Exception as e:
        log.warning("Download failed for %s: %s", url[:120], str(e)[:200])
        if fpath and fpath.exists():
            try:
                fpath.unlink(missing_ok=True)
            except Exception:
                pass
        return None


def _crawl_niin_on_sites(db, session, niin, desc, pns, sites, cfg, dl_cfg):
    """Core crawl logic for a single NIIN across all sites.
    Returns (results_count, files_count)."""
    results_count = 0
    files_count = 0
    max_size_bytes = int(dl_cfg["MAX_FILE_SIZE_MB"]) * 1024 * 1024
    max_files = int(dl_cfg["MAX_FILES_PER_NIIN"])
    min_relevance = float(dl_cfg["MIN_RELEVANCE_FOR_DOWNLOAD"])
    # Count existing files for this NIIN
    existing_files = db.execute("SELECT COUNT(*) c FROM crawler_files WHERE niin=?", (niin,)).fetchone()["c"]
    total_files = existing_files

    search_terms = [niin]
    if desc and desc.strip():
        search_terms.append(f"{niin} {desc.strip()}")
    for pn in pns[:2]:
        if pn.strip():
            search_terms.append(pn.strip())

    # Build search URL patterns per site
    SEARCH_PATTERNS = {
        "archive.org": "https://archive.org/search?query={term}",
        "dtic.mil": "{base}/search?q={term}",
        "everyspec.com": "{base}/search.php?query={term}",
    }

    import urllib.parse as _urlparse
    failed_sites = set()  # Skip sites that fail
    start_time = time.time()
    max_crawl_time = 180  # 3 minutes max per NIIN

    for site_row in sites:
        # Check overall time limit
        if time.time() - start_time > max_crawl_time:
            log.info("Crawler: time limit reached for NIIN %s after %ds", niin, int(time.time()-start_time))
            break

        site_url = site_row["url"]
        site_name = site_row["site_name"]
        base = site_url.rstrip("/")

        if site_name in failed_sites:
            continue

        # Determine search URL pattern for this site
        search_pattern = None
        domain = site_url.replace("https://", "").replace("http://", "").split("/")[0].lower()
        for pattern_domain, pattern in SEARCH_PATTERNS.items():
            if pattern_domain in domain:
                search_pattern = pattern
                break
        if not search_pattern:
            search_pattern = "{base}/search?q={term}"

        for term in search_terms[:3]:
            if site_name in failed_sites:
                break
            try:
                encoded_term = _urlparse.quote_plus(term)
                surl = search_pattern.format(base=base, term=encoded_term)
                try:
                    resp = session.get(surl, timeout=8, allow_redirects=True)
                    if resp.status_code != 200:
                        log.debug("Crawler: %s returned %d for term '%s'", site_name, resp.status_code, term[:50])
                        time.sleep(0.3)
                        continue
                    soup = _BS4(resp.text, "html.parser")
                    title = soup.title.string if soup.title else ""
                    text = soup.get_text(separator=" ", strip=True)[:2000]

                    found_pns = []
                    found_specs = []
                    relevance = 0.0
                    if niin in text:
                        relevance += 50
                    for pn in pns:
                        if pn.strip() and pn.strip().lower() in text.lower():
                            found_pns.append(pn.strip())
                            relevance += 25
                    for kw in ("datasheet", "specification", "technical data", "drawing",
                               "mil-spec", "pdf", "download", "catalog"):
                        if kw in text.lower():
                            found_specs.append(kw)
                            relevance += 5
                    if relevance > 0:
                        idx = text.lower().find(niin)
                        if idx < 0 and found_pns:
                            idx = text.lower().find(found_pns[0].lower())
                        snippet = text[max(0, idx - 100):max(0, idx - 100) + 300].strip() if idx >= 0 else text[:300].strip()

                        db.execute("""
                            INSERT INTO crawler_results
                            (niin, site_name, source_url, page_url, page_title,
                             snippet, found_pns, found_specs, relevance_score)
                            VALUES (?,?,?,?,?,?,?,?,?)
                        """, (niin, site_name, site_url, surl, title,
                              snippet, ",".join(found_pns), ",".join(found_specs),
                              min(100, relevance)))
                        result_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
                        results_count += 1
                        log.info("Crawler: found result for NIIN %s on %s (relevance=%d)", niin, site_name, relevance)

                        # Download files if relevance is high enough
                        if relevance >= min_relevance and total_files < max_files:
                            dl_links = _detect_downloadable_links(soup, surl)
                            log.info("Crawler: found %d downloadable links for NIIN %s on %s", len(dl_links), niin, site_name)
                            for link in dl_links:
                                if total_files >= max_files:
                                    break
                                ft = link["file_type"]
                                if ft == "pdf" and not int(dl_cfg["DOWNLOAD_PDFS"]):
                                    continue
                                if ft == "image" and not int(dl_cfg["DOWNLOAD_IMAGES"]):
                                    continue
                                if ft == "document" and not int(dl_cfg["DOWNLOAD_DOCS"]):
                                    continue
                                dl_result = _download_crawler_file(session, link["url"], niin,
                                                                   max_size_bytes, CRAWLER_DL_DIR)
                                if dl_result:
                                    db.execute("""
                                        INSERT INTO crawler_files (niin, crawler_result_id, filename,
                                            file_path, file_type, file_size, mime_type, source_url)
                                        VALUES (?,?,?,?,?,?,?,?)
                                    """, (niin, result_id, dl_result["filename"], dl_result["file_path"],
                                          ft, dl_result["file_size"],
                                          dl_result["mime_type"], link["url"]))
                                    total_files += 1
                                    files_count += 1
                except _requests_lib.RequestException as re:
                    log.debug("Crawler: request error on %s for '%s': %s", site_name, term[:30], str(re)[:100])
                    failed_sites.add(site_name)  # Skip this site for remaining terms
                time.sleep(0.3)
            except Exception as ex:
                log.warning("Crawler: unexpected error on %s for NIIN %s: %s", site_name, niin, str(ex)[:200])
                failed_sites.add(site_name)
                continue

    elapsed = int(time.time() - start_time)
    log.info("Crawler: NIIN %s completed in %ds — %d results, %d files, %d sites failed",
             niin, elapsed, results_count, files_count, len(failed_sites))
    return results_count, files_count


def run_crawler_single_niin_worker(log_id, niin):
    """Crawl enabled sites for a single specific NIIN."""
    try:
        db = _worker_db()
        # Use top 15 sites for single-NIIN crawl (faster than all 47+)
        sites = db.execute("SELECT * FROM crawler_sites WHERE enabled=1 ORDER BY rank LIMIT 15").fetchall()
        if not sites:
            _update_log(log_id, "success", 0, "No crawler sites configured.")
            db.close()
            return
        # Get part numbers for this NIIN
        refs = db.execute("SELECT DISTINCT part_number FROM publog_references WHERE niin=?", (niin,)).fetchall()
        pns = [r["part_number"] for r in refs if r["part_number"]]
        desc_row = db.execute("SELECT item_name FROM publog_identification WHERE niin=?", (niin,)).fetchone()
        desc = desc_row["item_name"] if desc_row else ""
        dl_cfg = load_config_from_db(db, "crawler", CRAWLER_DEFAULTS)
        session = _requests_lib.Session()
        session.headers.update({"User-Agent": "Mozilla/5.0 (compatible; NSNIntelBot/1.0; research)",
                                "Accept": "text/html,application/xhtml+xml"})
        log.info("Starting single-NIIN crawl for %s across %d sites", niin, len(sites))
        r_cnt, f_cnt = _crawl_niin_on_sites(db, session, niin, desc, pns, sites, {}, dl_cfg)
        db.commit()
        db.close()
        _update_log(log_id, "success", r_cnt,
                    f"Crawled {len(sites)} sites for NIIN {niin}. {r_cnt} results, {f_cnt} files downloaded.")
    except Exception as e:
        log.exception("Single-NIIN crawler failed for %s", niin)
        _update_log(log_id, "error", log_text=str(e))


def run_crawler_batch_worker(log_id, niins):
    """Crawl enabled sites for a list of NIINs."""
    try:
        db = _worker_db()
        sites = db.execute("SELECT * FROM crawler_sites WHERE enabled=1 ORDER BY rank").fetchall()
        if not sites:
            _update_log(log_id, "success", 0, "No crawler sites configured.")
            db.close()
            return
        dl_cfg = load_config_from_db(db, "crawler", CRAWLER_DEFAULTS)
        session = _requests_lib.Session()
        session.headers.update({"User-Agent": "Mozilla/5.0 (compatible; NSNIntelBot/1.0; research)",
                                "Accept": "text/html,application/xhtml+xml"})
        total_r, total_f = 0, 0
        for niin in niins:
            refs = db.execute("SELECT DISTINCT part_number FROM publog_references WHERE niin=?", (niin,)).fetchall()
            pns = [r["part_number"] for r in refs if r["part_number"]]
            desc_row = db.execute("SELECT item_name FROM publog_identification WHERE niin=?", (niin,)).fetchone()
            desc = desc_row["item_name"] if desc_row else ""
            r_cnt, f_cnt = _crawl_niin_on_sites(db, session, niin, desc, pns, sites, {}, dl_cfg)
            total_r += r_cnt
            total_f += f_cnt
            db.commit()
        db.close()
        _update_log(log_id, "success", total_r,
                    f"Batch crawled {len(niins)} NIINs. {total_r} results, {total_f} files.")
    except Exception as e:
        log.exception("Batch crawler failed")
        _update_log(log_id, "error", log_text=str(e))


@app.route("/api/run/crawler-niin", methods=["POST"])
def api_run_crawler_niin():
    data = request.json if request.is_json else {}
    niin = data.get("niin")
    if not niin:
        return jsonify({"error": "niin required"}), 400
    log_id = _start_pipeline("crawler_niin", run_crawler_single_niin_worker, niin)
    return jsonify({"job_id": log_id, "status": "running"})


@app.route("/api/run/crawler-batch", methods=["POST"])
def api_run_crawler_batch():
    data = request.json if request.is_json else {}
    niins = data.get("niins", [])
    if not niins:
        return jsonify({"error": "niins list required"}), 400
    log_id = _start_pipeline("crawler_batch", run_crawler_batch_worker, niins)
    return jsonify({"job_id": log_id, "status": "running"})


@app.route("/api/crawler/files")
def api_crawler_files():
    db = get_db()
    niin = request.args.get("niin", "")
    where, params = ["1=1"], []
    if niin:
        where.append("cf.niin = ?")
        params.append(niin)
    rows = db.execute(f"""
        SELECT cf.*, cr.page_title, cr.site_name, cr.relevance_score
        FROM crawler_files cf
        LEFT JOIN crawler_results cr ON cf.crawler_result_id = cr.id
        WHERE {" AND ".join(where)}
        ORDER BY cf.downloaded_at DESC
    """, params).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/crawler/files/<int:file_id>/download")
def api_crawler_file_download(file_id):
    db = get_db()
    row = db.execute("SELECT * FROM crawler_files WHERE id=?", (file_id,)).fetchone()
    db.close()
    if not row:
        return jsonify({"error": "not found"}), 404
    fpath = Path(row["file_path"])
    if not fpath.exists():
        return jsonify({"error": "file missing from disk"}), 404
    return send_file(str(fpath), as_attachment=True, download_name=row["filename"])


# ──────────────────────────────────────────────
# SECTION I-B-3: Document Intelligence
# ──────────────────────────────────────────────

def _extract_from_pdf(file_path):
    """Extract text and tables from a PDF file using pdfplumber."""
    extracts = []
    try:
        import pdfplumber
        with pdfplumber.open(file_path) as pdf:
            all_text = []
            for page in pdf.pages[:20]:
                text = page.extract_text() or ""
                all_text.append(text)
                tables = page.extract_tables()
                for table in tables:
                    table_text = "\n".join([" | ".join(str(c or "") for c in row) for row in table])
                    extracts.append({
                        "extract_type": "table",
                        "content": table_text[:5000],
                        "niin_mentions": ",".join(set(NIIN_PATTERN.findall(table_text) + NSN_PATTERN.findall(table_text))),
                        "pn_mentions": ",".join(list(set(PN_PATTERN.findall(table_text)))[:20]),
                    })
            full_text = "\n".join(all_text)
            if full_text.strip():
                all_niins = set(NIIN_PATTERN.findall(full_text)) | set(NIIN_PATTERN.findall(
                    "".join(m.replace("-", "") for m in NSN_PATTERN.findall(full_text))))
                extracts.append({
                    "extract_type": "text",
                    "content": full_text[:10000],
                    "niin_mentions": ",".join(set(NIIN_PATTERN.findall(full_text) + NSN_PATTERN.findall(full_text))),
                    "pn_mentions": ",".join(list(set(PN_PATTERN.findall(full_text)))[:30]),
                })
    except Exception as e:
        extracts.append({"extract_type": "text", "content": f"Extraction error: {e}",
                         "niin_mentions": "", "pn_mentions": ""})
    return extracts


def _extract_from_image(file_path):
    """Extract metadata from an image file."""
    try:
        size = Path(file_path).stat().st_size
        return [{"extract_type": "image_ref",
                 "content": json.dumps({"file_size": size, "path": file_path}),
                 "niin_mentions": "", "pn_mentions": ""}]
    except Exception as e:
        return [{"extract_type": "image_ref", "content": f"Error: {e}",
                 "niin_mentions": "", "pn_mentions": ""}]


def run_extract_file_worker(log_id, file_id):
    """Extract information from a single crawler file."""
    try:
        db = _worker_db()
        row = db.execute("SELECT * FROM crawler_files WHERE id=?", (file_id,)).fetchone()
        if not row:
            _update_log(log_id, "error", log_text=f"File ID {file_id} not found")
            db.close()
            return
        ftype = row["file_type"]
        if ftype in ("pdf", "datasheet"):
            extracts = _extract_from_pdf(row["file_path"])
        elif ftype == "image":
            extracts = _extract_from_image(row["file_path"])
        else:
            extracts = [{"extract_type": "text", "content": f"Unsupported file type: {ftype}",
                         "niin_mentions": "", "pn_mentions": ""}]
        count = 0
        for ext in extracts:
            db.execute("""INSERT INTO crawler_file_extracts (file_id, extract_type, content, niin_mentions, pn_mentions)
                VALUES (?,?,?,?,?)""", (file_id, ext["extract_type"], ext["content"],
                                        ext["niin_mentions"], ext["pn_mentions"]))
            count += 1
        db.commit()
        db.close()
        _update_log(log_id, "success", count, f"Extracted {count} segments from file {file_id}")
    except Exception as e:
        log.exception("File extraction failed for file %s", file_id)
        _update_log(log_id, "error", log_text=str(e))


def run_extract_all_worker(log_id, niin):
    """Extract from all unprocessed crawler files for a NIIN."""
    try:
        db = _worker_db()
        files = db.execute("""
            SELECT cf.id, cf.file_path, cf.file_type
            FROM crawler_files cf
            LEFT JOIN crawler_file_extracts cfe ON cf.id = cfe.file_id
            WHERE cf.niin = ? AND cfe.id IS NULL
        """, (niin,)).fetchall()
        total = 0
        for f in files:
            ftype = f["file_type"]
            if ftype in ("pdf", "datasheet"):
                extracts = _extract_from_pdf(f["file_path"])
            elif ftype == "image":
                extracts = _extract_from_image(f["file_path"])
            else:
                extracts = []
            for ext in extracts:
                db.execute("""INSERT INTO crawler_file_extracts (file_id, extract_type, content, niin_mentions, pn_mentions)
                    VALUES (?,?,?,?,?)""", (f["id"], ext["extract_type"], ext["content"],
                                            ext["niin_mentions"], ext["pn_mentions"]))
                total += 1
        db.commit()
        db.close()
        _update_log(log_id, "success", total, f"Extracted {total} segments from {len(files)} files for NIIN {niin}")
    except Exception as e:
        log.exception("Batch extraction failed for NIIN %s", niin)
        _update_log(log_id, "error", log_text=str(e))


@app.route("/api/crawler/files/<int:file_id>/extract", methods=["POST"])
def api_extract_file(file_id):
    log_id = _start_pipeline("extract_file", run_extract_file_worker, file_id)
    return jsonify({"job_id": log_id, "status": "running"})


@app.route("/api/crawler/extract-all", methods=["POST"])
def api_extract_all():
    data = request.json if request.is_json else {}
    niin = data.get("niin") or request.args.get("niin")
    if not niin:
        return jsonify({"error": "niin required"}), 400
    log_id = _start_pipeline("extract_all", run_extract_all_worker, niin)
    return jsonify({"job_id": log_id, "status": "running"})


@app.route("/api/crawler/files/<int:file_id>/extracts")
def api_file_extracts(file_id):
    db = get_db()
    rows = db.execute("SELECT * FROM crawler_file_extracts WHERE file_id=? ORDER BY id", (file_id,)).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/crawler/extracts")
def api_crawler_extracts():
    db = get_db()
    niin = request.args.get("niin", "")
    if not niin:
        db.close()
        return jsonify([])
    rows = db.execute("""
        SELECT cfe.*, cf.filename, cf.file_type, cf.source_url
        FROM crawler_file_extracts cfe
        JOIN crawler_files cf ON cfe.file_id = cf.id
        WHERE cf.niin = ?
        ORDER BY cfe.extracted_at DESC
    """, (niin,)).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


# ──────────────────────────────────────────────
# SECTION I-D: Gemini AI Image Generation
# ──────────────────────────────────────────────

def generate_part_image(niin, spec):
    """Generate a photorealistic image using Google Gemini Imagen API."""
    import base64
    from google import genai
    from google.genai import types as genai_types

    db = get_db()
    api_key_row = db.execute(
        "SELECT value FROM config WHERE key='GEMINI_API_KEY' AND section='general'"
    ).fetchone()
    db.close()
    if not api_key_row or not api_key_row["value"]:
        raise ValueError("GEMINI_API_KEY not configured. Set it in Configuration > General.")

    client = genai.Client(api_key=api_key_row["value"])

    item_name = spec.get("item_name", "industrial part")
    materials = ", ".join(spec.get("materials", ["metal"])[:3])
    features = ", ".join(f.split("(")[0].strip() for f in spec.get("features", [])[:5])
    dims = spec.get("dimensions", {})
    dim_desc = ", ".join(f"{k.replace('_', ' ')}: {v}mm" for k, v in list(dims.items())[:5])

    prompt = (
        f"Photorealistic studio photograph of a {item_name}. "
        f"Material: {materials}. Key features: {features}. "
        f"Approximate dimensions: {dim_desc}. "
        f"Industrial/engineering part on a clean white background, "
        f"professional product photography, high detail, sharp focus, no text."
    )

    response = client.models.generate_images(
        model="imagen-4.0-generate-001",
        prompt=prompt,
        config=genai_types.GenerateImagesConfig(
            number_of_images=1,
            output_mime_type="image/png",
        ),
    )

    if not response.generated_images:
        raise ValueError("No image generated (may have been filtered by safety)")

    img = response.generated_images[0].image
    out_dir = CAD_OUTPUT_DIR / f"NIIN_{niin}"
    out_dir.mkdir(parents=True, exist_ok=True)
    img_path = out_dir / f"ai_image_{int(time.time())}.png"
    img.save(str(img_path))

    b64 = base64.b64encode(img_path.read_bytes()).decode("utf-8")
    return {"image_path": str(img_path), "base64": b64, "prompt": prompt}


def run_image_generation_worker(log_id, niin):
    """Background worker for AI image generation."""
    try:
        log.info("Starting AI image generation for NIIN %s", niin)
        db = _worker_db()
        spec = build_part_spec_from_publog(db, niin)
        db.close()
        log.info("Built spec for NIIN %s: type=%s, name=%s", niin,
                 spec.get("part_type", "?"), spec.get("item_name", "?"))

        result = generate_part_image(niin, spec)
        log.info("Image generated for NIIN %s, saving to DB...", niin)

        db = _worker_db()
        db.execute("""
            INSERT INTO cad_jobs (niin, status, spec_json, output_dir, image_path, job_type, finished_at)
            VALUES (?, 'completed', ?, ?, ?, 'ai_image', datetime('now'))
        """, (niin, json.dumps({"type": "ai_image", "prompt": result["prompt"]}),
              str(CAD_OUTPUT_DIR / f"NIIN_{niin}"), result["image_path"]))
        db.commit()
        db.close()
        _update_log(log_id, "success", 1, f"Generated AI image for NIIN {niin}")
    except Exception as e:
        log.exception("Image generation failed for NIIN %s", niin)
        _update_log(log_id, "error", log_text=str(e)[:500])


@app.route("/api/niin/<niin>/generate-image", methods=["POST"])
def api_generate_image(niin):
    log_id = _start_pipeline("generate_image", run_image_generation_worker, niin)
    return jsonify({"job_id": log_id, "status": "running"})


@app.route("/api/cad/image/<niin>")
def api_cad_image(niin):
    """Serve the most recent AI-generated image for a NIIN."""
    db = get_db()
    row = db.execute("""
        SELECT image_path FROM cad_jobs
        WHERE niin=? AND image_path != '' AND image_path IS NOT NULL
        ORDER BY finished_at DESC LIMIT 1
    """, (niin,)).fetchone()
    db.close()
    if not row or not row["image_path"]:
        return jsonify({"error": "No image found"}), 404
    fpath = Path(row["image_path"])
    if not fpath.exists():
        return jsonify({"error": "Image file missing from disk"}), 404
    return send_file(str(fpath), mimetype="image/png")


# ──────────────────────────────────────────────
# SECTION I-E: DXF Engineering Drawing Generation
# ──────────────────────────────────────────────

def generate_dxf(niin, spec):
    """Generate a 2D engineering DXF drawing from part specification."""
    import ezdxf

    dims = spec.get("dimensions", {})
    part_type = spec.get("part_type", "cylinder")
    item_name = spec.get("item_name", "")
    materials = ", ".join(spec.get("materials", []))

    doc = ezdxf.new("R2010")
    msp = doc.modelspace()

    # Title block (bottom-right)
    _draw_title_block(msp, niin, item_name, materials, spec.get("units", "mm"))

    if part_type == "flange" or "flange_od" in dims:
        _draw_flange_views(msp, dims)
    elif part_type == "block" or all(k in dims for k in ("length", "width", "height")):
        _draw_block_views(msp, dims)
    else:
        _draw_cylinder_views(msp, dims)

    out_dir = CAD_OUTPUT_DIR / f"NIIN_{niin}"
    out_dir.mkdir(parents=True, exist_ok=True)
    dxf_path = out_dir / f"drawing_{int(time.time())}.dxf"
    doc.saveas(str(dxf_path))
    return {"dxf_path": str(dxf_path)}


def _draw_title_block(msp, niin, item_name, materials, units):
    """Draw a standard title block."""
    x0, y0 = 200, -20
    pts = [(x0, y0), (x0 + 180, y0), (x0 + 180, y0 - 60), (x0, y0 - 60), (x0, y0)]
    for i in range(4):
        msp.add_line(pts[i], pts[i + 1])
    msp.add_line((x0, y0 - 15), (x0 + 180, y0 - 15))  # divider
    msp.add_line((x0, y0 - 30), (x0 + 180, y0 - 30))
    msp.add_line((x0, y0 - 45), (x0 + 180, y0 - 45))
    msp.add_text(f"NIIN: {niin}", dxfattribs={"height": 3.5}).set_placement((x0 + 5, y0 - 12))
    msp.add_text(f"ITEM: {item_name[:50]}", dxfattribs={"height": 2.5}).set_placement((x0 + 5, y0 - 24))
    msp.add_text(f"MATERIAL: {materials[:50]}", dxfattribs={"height": 2}).set_placement((x0 + 5, y0 - 38))
    msp.add_text(f"UNITS: {units}", dxfattribs={"height": 2}).set_placement((x0 + 100, y0 - 38))
    msp.add_text("NSN Intelligence Suite - Auto Generated", dxfattribs={"height": 1.8}).set_placement((x0 + 5, y0 - 55))


def _draw_flange_views(msp, dims):
    """Draw flange: top view (circle + bolt pattern) and front view (cross section)."""
    od = float(dims.get("flange_od", 120))
    thk = float(dims.get("flange_thickness", 15))
    bore = float(dims.get("center_bore", 40))
    bcd = float(dims.get("bolt_circle_diameter", od * 0.75))
    bolt_qty = int(dims.get("bolt_hole_qty", 4))
    bolt_dia = float(dims.get("bolt_hole_diameter", od * 0.08))

    # TOP VIEW centered at (80, 80)
    cx, cy = 80, 80
    msp.add_circle((cx, cy), od / 2)
    if bore > 0:
        msp.add_circle((cx, cy), bore / 2)
    if bolt_qty > 0:
        msp.add_circle((cx, cy), bcd / 2)
        for i in range(bolt_qty):
            angle = 2 * math.pi * i / bolt_qty
            bx = cx + (bcd / 2) * math.cos(angle)
            by = cy + (bcd / 2) * math.sin(angle)
            msp.add_circle((bx, by), bolt_dia / 2)
    # Dimension labels
    msp.add_text(f"OD {od}", dxfattribs={"height": 2.5}).set_placement((cx + od / 2 + 3, cy + 2))
    if bore > 0:
        msp.add_text(f"BORE {bore}", dxfattribs={"height": 2}).set_placement((cx + 2, cy - bore / 2 - 6))
    if bolt_qty > 0:
        msp.add_text(f"BCD {bcd}", dxfattribs={"height": 1.8}).set_placement((cx + bcd / 2 + 3, cy - 5))
        msp.add_text(f"{bolt_qty}x {bolt_dia} HOLES", dxfattribs={"height": 1.8}).set_placement((cx - od / 2 - 5, cy + od / 2 + 5))
    msp.add_text("TOP VIEW", dxfattribs={"height": 3}).set_placement((cx - 15, cy - od / 2 - 12))

    # FRONT VIEW at (250, 80)
    fx, fy = 250, 80
    # Outer rectangle
    pts = [(fx, fy - od / 2), (fx + thk, fy - od / 2), (fx + thk, fy + od / 2), (fx, fy + od / 2)]
    for i in range(4):
        msp.add_line(pts[i], pts[(i + 1) % 4])
    if bore > 0:
        msp.add_line((fx, fy - bore / 2), (fx + thk, fy - bore / 2))
        msp.add_line((fx, fy + bore / 2), (fx + thk, fy + bore / 2))
    msp.add_text(f"THK {thk}", dxfattribs={"height": 2}).set_placement((fx + thk / 2 - 5, fy - od / 2 - 8))
    msp.add_text(f"OD {od}", dxfattribs={"height": 2}).set_placement((fx + thk + 3, fy))
    msp.add_text("FRONT VIEW", dxfattribs={"height": 3}).set_placement((fx - 5, fy - od / 2 - 16))


def _draw_block_views(msp, dims):
    """Draw block: front (L×H), side (W×H), top (L×W)."""
    L = float(dims.get("length", 100))
    W = float(dims.get("width", 60))
    H = float(dims.get("height", 30))

    # FRONT VIEW at (80, 60)
    x0 = 80 - L / 2
    y0 = 60 - H / 2
    pts = [(x0, y0), (x0 + L, y0), (x0 + L, y0 + H), (x0, y0 + H)]
    for i in range(4):
        msp.add_line(pts[i], pts[(i + 1) % 4])
    msp.add_text(f"{L}", dxfattribs={"height": 2.5}).set_placement((80, y0 - 6))
    msp.add_text(f"{H}", dxfattribs={"height": 2.5}).set_placement((x0 - 10, 60))
    msp.add_text("FRONT", dxfattribs={"height": 3}).set_placement((80 - 8, y0 - 14))

    # SIDE VIEW at (220, 60)
    sx = 220
    s0 = sx - W / 2
    pts2 = [(s0, y0), (s0 + W, y0), (s0 + W, y0 + H), (s0, y0 + H)]
    for i in range(4):
        msp.add_line(pts2[i], pts2[(i + 1) % 4])
    msp.add_text(f"{W}", dxfattribs={"height": 2.5}).set_placement((sx, y0 - 6))
    msp.add_text("SIDE", dxfattribs={"height": 3}).set_placement((sx - 6, y0 - 14))

    # TOP VIEW at (80, 140)
    ty = 140
    t0x = 80 - L / 2
    t0y = ty - W / 2
    pts3 = [(t0x, t0y), (t0x + L, t0y), (t0x + L, t0y + W), (t0x, t0y + W)]
    for i in range(4):
        msp.add_line(pts3[i], pts3[(i + 1) % 4])
    msp.add_text("TOP", dxfattribs={"height": 3}).set_placement((80 - 5, t0y - 8))

    # Holes in top view
    hole_qty = int(dims.get("hole_qty", 0))
    hole_d = float(dims.get("hole_diameter", 8))
    spacing = float(dims.get("hole_spacing", L * 0.5))
    if hole_qty > 0:
        for i in range(hole_qty):
            if hole_qty <= 1:
                hx = 80
            else:
                hx = 80 - spacing / 2 + i * (spacing / max(hole_qty - 1, 1))
            msp.add_circle((hx, ty), hole_d / 2)


def _draw_cylinder_views(msp, dims):
    """Draw cylinder: front view (rectangle) and end view (circle)."""
    od = float(dims.get("flange_od", dims.get("od", 50)))
    length = float(dims.get("flange_thickness", dims.get("length", 100)))
    bore = float(dims.get("center_bore", dims.get("bore", 0)))

    # FRONT VIEW at (80, 80)
    x0 = 80 - length / 2
    y0 = 80 - od / 2
    pts = [(x0, y0), (x0 + length, y0), (x0 + length, y0 + od), (x0, y0 + od)]
    for i in range(4):
        msp.add_line(pts[i], pts[(i + 1) % 4])
    msp.add_text(f"L={length}", dxfattribs={"height": 2.5}).set_placement((80, y0 - 6))
    msp.add_text(f"OD={od}", dxfattribs={"height": 2.5}).set_placement((x0 - 18, 80))
    msp.add_text("FRONT", dxfattribs={"height": 3}).set_placement((80 - 8, y0 - 14))

    # END VIEW at (220, 80)
    cx2 = 220
    msp.add_circle((cx2, 80), od / 2)
    if bore > 0:
        msp.add_circle((cx2, 80), bore / 2)
    msp.add_text("END", dxfattribs={"height": 3}).set_placement((cx2 - 5, 80 - od / 2 - 14))


def run_dxf_generation_worker(log_id, niin):
    """Background worker for DXF generation."""
    try:
        db = _worker_db()
        spec = build_part_spec_from_publog(db, niin)
        result = generate_dxf(niin, spec)
        db.execute("""
            INSERT INTO cad_jobs (niin, status, spec_json, output_dir, dxf_path, job_type, finished_at)
            VALUES (?, 'completed', ?, ?, ?, 'dxf', datetime('now'))
        """, (niin, json.dumps(spec, default=str), str(CAD_OUTPUT_DIR / f"NIIN_{niin}"), result["dxf_path"]))
        db.commit()
        db.close()
        _update_log(log_id, "success", 1, f"Generated DXF drawing for NIIN {niin}")
    except Exception as e:
        log.exception("DXF generation failed for NIIN %s", niin)
        _update_log(log_id, "error", log_text=str(e))


@app.route("/api/niin/<niin>/generate-dxf", methods=["POST"])
def api_generate_dxf(niin):
    log_id = _start_pipeline("generate_dxf", run_dxf_generation_worker, niin)
    return jsonify({"job_id": log_id, "status": "running"})


@app.route("/api/cad/dxf/<niin>")
def api_download_dxf(niin):
    db = get_db()
    row = db.execute("""
        SELECT dxf_path FROM cad_jobs
        WHERE niin=? AND dxf_path != '' AND dxf_path IS NOT NULL
        ORDER BY finished_at DESC LIMIT 1
    """, (niin,)).fetchone()
    db.close()
    if not row or not row["dxf_path"]:
        return jsonify({"error": "No DXF file found"}), 404
    fpath = Path(row["dxf_path"])
    if not fpath.exists():
        return jsonify({"error": "DXF file missing from disk"}), 404
    return send_file(str(fpath), as_attachment=True, download_name=f"NIIN_{niin}.dxf")


# ──────────────────────────────────────────────
# SECTION I-C: Text-to-CAD Feature
# ──────────────────────────────────────────────

CAD_OUTPUT_DIR = BASE_DIR / "cad_output"

def _classify_part_type(item_name: str, features: list, dimensions: dict) -> str:
    """Classify part into geometry type: 'flange', 'block', or 'cylinder' based on PUBLOG data."""
    name_lower = (item_name or "").lower()
    feat_lower = " ".join(f.lower() for f in features)
    dim_keys = " ".join(k.lower() for k in dimensions.keys())

    # Flange detection: explicit flange keywords or bolt pattern + circular features
    if ("flange" in name_lower or "flange" in feat_lower or
        "bolt_circle" in dim_keys or "bolt_hole" in dim_keys or
        ("flange" in dim_keys)):
        return "flange"

    # Block detection: rectangular keywords or length+width+height present
    block_keys = {"length", "width", "height"}
    if (any(k in name_lower for k in ("block", "plate", "bracket", "mount", "bar", "panel")) or
        any(k in feat_lower for k in ("rectangular", "block", "plate")) or
        block_keys.issubset(set(k.lower() for k in dimensions.keys()))):
        return "block"

    # Default to cylinder (most common for valves, fittings, bushings, etc.)
    return "cylinder"


def _extract_flange_dims(dimensions: dict, raw_dims: dict) -> dict:
    """Attempt to map extracted dimensions to flange-specific parameters."""
    result = {}
    mapped_keys = set()

    # Direct mapping of known keys
    key_map = {
        "flange_od": ["flange_od", "flange_diameter", "flange_outside_diameter"],
        "flange_thickness": ["flange_thickness", "flange_thk", "flange_t", "flange_width"],
        "bolt_circle_diameter": ["bolt_circle_diameter", "bcd", "bolt_circle"],
        "bolt_hole_qty": ["bolt_hole_qty", "bolt_holes", "bolt_quantity", "number_of_bolt_holes"],
        "bolt_hole_diameter": ["bolt_hole_diameter", "bolt_hole_dia", "bolt_dia"],
        "center_bore": ["center_bore", "bore", "bore_diameter", "id", "inside_diameter", "inner_diameter"],
        "raised_face_diameter": ["raised_face_diameter", "rf_diameter", "raised_face_dia"],
        "raised_face_height": ["raised_face_height", "rf_height"],
    }
    for target, sources in key_map.items():
        for src in sources:
            for dk, dv in dimensions.items():
                if src in dk.lower().replace(" ", "_"):
                    result[target] = dv
                    mapped_keys.add(dk)
                    break
            if target in result:
                break

    # If no explicit flange_od, try to find the largest OD-like dimension
    if "flange_od" not in result:
        od_candidates = [(k, v) for k, v in dimensions.items()
                         if any(x in k.lower() for x in ("od", "outside", "outer", "diameter", "diam"))
                         and k not in mapped_keys]
        if od_candidates:
            od_candidates.sort(key=lambda x: x[1], reverse=True)
            result["flange_od"] = od_candidates[0][1]
            mapped_keys.add(od_candidates[0][0])

    # If no flange_thickness, try thickness/height-like dims
    if "flange_thickness" not in result:
        thk_candidates = [(k, v) for k, v in dimensions.items()
                          if any(x in k.lower() for x in ("thickness", "thk", "height", "width"))
                          and k not in mapped_keys]
        if thk_candidates:
            thk_candidates.sort(key=lambda x: x[1])  # smallest is likely thickness
            result["flange_thickness"] = thk_candidates[0][1]
            mapped_keys.add(thk_candidates[0][0])

    # If no center_bore, try bore/id dims
    if "center_bore" not in result:
        bore_candidates = [(k, v) for k, v in dimensions.items()
                           if any(x in k.lower() for x in ("bore", "id", "inside", "inner"))
                           and k not in mapped_keys]
        if bore_candidates:
            result["center_bore"] = bore_candidates[0][1]
            mapped_keys.add(bore_candidates[0][0])

    return result


def _extract_block_dims(dimensions: dict) -> dict:
    """Attempt to map extracted dimensions to block-specific parameters."""
    result = {}
    for dk, dv in dimensions.items():
        kl = dk.lower()
        if "length" in kl or "long" in kl:
            result.setdefault("length", dv)
        elif "width" in kl or "wide" in kl:
            result.setdefault("width", dv)
        elif "height" in kl or "thick" in kl or "depth" in kl:
            result.setdefault("height", dv)
        elif "hole_dia" in kl or "hole_d" in kl:
            result.setdefault("hole_diameter", dv)
        elif "hole_spacing" in kl or "hole_space" in kl:
            result.setdefault("hole_spacing", dv)
        elif "hole_qty" in kl or "hole_count" in kl or "number_of_holes" in kl:
            result.setdefault("hole_qty", int(dv))
    return result


def build_part_spec_from_publog(db, niin: str) -> dict:
    """Convert PUBLOG characteristics into a part specification JSON for CAD generation.
    Uses v2 geometry types: flange-like (with bolt patterns) and block-like (with holes).
    """
    ident = db.execute("SELECT * FROM publog_identification WHERE niin=?", (niin,)).fetchone()
    chars = db.execute("""
        SELECT mrc, requirements_statement, clear_text_reply
        FROM publog_characteristics WHERE niin=? ORDER BY mrc
    """, (niin,)).fetchall()

    part_name = f"NIIN_{niin}"
    item_name = ident["item_name"] if ident else "Unknown"

    # Parse dimensions and features from characteristics
    raw_dimensions = {}  # All extracted numeric dims
    features = []
    materials = []
    tolerances = "General +/- 0.1mm unless noted"

    for c in chars:
        stmt = (c["requirements_statement"] or "").strip()
        reply = (c["clear_text_reply"] or "").strip()
        mrc = (c["mrc"] or "").strip()

        if not reply:
            continue

        combined = f"{stmt} {reply}".lower()

        # Try to extract dimensional values
        dim_match = re.search(r'(\d+\.?\d*)\s*(mm|in|inch|inches|cm|")', combined)
        if dim_match:
            val = float(dim_match.group(1))
            unit = dim_match.group(2).replace('"', 'in').replace('inches', 'in').replace('inch', 'in')
            if unit == 'in':
                val = val * 25.4
            elif unit == 'cm':
                val = val * 10
            dim_name = re.sub(r'[^a-z0-9_]', '_', stmt.lower())[:30] if stmt else f"dim_{mrc}"
            raw_dimensions[dim_name] = round(val, 2)

        # Detect features
        for feat_kw in ("bore", "thread", "flange", "seal", "port", "groove", "chamfer",
                        "radius", "hole", "slot", "keyway", "o-ring", "bolt"):
            if feat_kw in combined:
                features.append(f"{feat_kw.title()} ({stmt})")

        # Detect materials
        for mat_kw in ("stainless", "steel", "aluminum", "brass", "bronze", "titanium",
                       "inconel", "copper", "rubber", "teflon", "ptfe", "nylon"):
            if mat_kw in combined:
                materials.append(reply)
                break

    # Classify part type and build appropriate dimensions
    part_type = _classify_part_type(item_name, features, raw_dimensions)

    if part_type == "flange":
        mapped = _extract_flange_dims(raw_dimensions, raw_dimensions)
        # Apply sensible defaults if we have at least an OD
        flange_od = mapped.get("flange_od", 120)
        dimensions = {
            "flange_od": flange_od,
            "flange_thickness": mapped.get("flange_thickness", round(flange_od * 0.12, 1)),
            "bolt_circle_diameter": mapped.get("bolt_circle_diameter", round(flange_od * 0.75, 1)),
            "bolt_hole_qty": int(mapped.get("bolt_hole_qty", 4)),
            "bolt_hole_diameter": mapped.get("bolt_hole_diameter", round(flange_od * 0.08, 1)),
            "center_bore": mapped.get("center_bore", round(flange_od * 0.33, 1)),
        }
        if "raised_face_diameter" in mapped:
            dimensions["raised_face_diameter"] = mapped["raised_face_diameter"]
            dimensions["raised_face_height"] = mapped.get("raised_face_height", 2)
    elif part_type == "block":
        mapped = _extract_block_dims(raw_dimensions)
        length = mapped.get("length", 100)
        dimensions = {
            "length": length,
            "width": mapped.get("width", round(length * 0.6, 1)),
            "height": mapped.get("height", round(length * 0.3, 1)),
            "hole_qty": int(mapped.get("hole_qty", 2)),
            "hole_diameter": mapped.get("hole_diameter", 8),
            "hole_spacing": mapped.get("hole_spacing", round(length * 0.5, 1)),
        }
    else:
        # Cylinder fallback — map to flange-like with simple body
        od_vals = [(k, v) for k, v in raw_dimensions.items()
                   if any(x in k.lower() for x in ("od", "outside", "outer", "diam"))]
        len_vals = [(k, v) for k, v in raw_dimensions.items()
                    if any(x in k.lower() for x in ("length", "long", "height", "body"))]
        bore_vals = [(k, v) for k, v in raw_dimensions.items()
                     if any(x in k.lower() for x in ("bore", "id", "inside", "inner"))]

        od = od_vals[0][1] if od_vals else 50
        length = len_vals[0][1] if len_vals else od * 2
        bore_d = bore_vals[0][1] if bore_vals else round(od * 0.35, 1)
        has_bore = any("bore" in f.lower() for f in features)
        has_flange = any("flange" in f.lower() for f in features)

        # Build as flange-like if it has any flange indicators
        if has_flange:
            dimensions = {
                "flange_od": round(od * 1.5, 1),
                "flange_thickness": round(length * 0.15, 1),
                "bolt_circle_diameter": round(od * 1.12, 1),
                "bolt_hole_qty": 4,
                "bolt_hole_diameter": round(od * 0.08, 1),
                "center_bore": bore_d if has_bore else round(od * 0.33, 1),
            }
        else:
            # Simple cylinder represented as flange with no bolt holes
            dimensions = {
                "flange_od": od,
                "flange_thickness": length,
                "center_bore": bore_d if has_bore else 0,
                "bolt_hole_qty": 0,
            }

    spec = {
        "part_name": part_name,
        "item_name": item_name,
        "niin": niin,
        "units": "mm",
        "part_type": part_type,
        "dimensions": dimensions,
        "raw_dimensions": raw_dimensions,
        "features": features if features else [item_name],
        "materials": materials if materials else ["Not specified"],
        "tolerances": tolerances,
        "characteristics_count": len(chars),
    }
    return spec


def generate_openscad(spec: dict) -> str:
    """Generate an OpenSCAD .scad file from a part specification (v2 geometry engine).
    Supports flange-like (with bolt patterns, raised faces) and block-like geometries.
    """
    dims = spec.get("dimensions", {})
    features = spec.get("features", [])
    name = spec.get("part_name", "part")
    part_type = spec.get("part_type", "cylinder")

    header = (
        f"// OpenSCAD model for {name}\n"
        f"// Item: {spec.get('item_name', '')}\n"
        f"// NIIN: {spec.get('niin', '')}\n"
        f"// Part type: {part_type}\n"
        f"// Generated by NSN Intelligence Suite (v2 engine)\n"
        f"// Characteristics: {spec.get('characteristics_count', 0)} PUBLOG entries\n"
        f"// Geometry units: mm\n"
    )

    # ── Flange-like geometry ──
    if "flange_od" in dims and "flange_thickness" in dims:
        flange_od = float(dims["flange_od"])
        flange_thk = float(dims["flange_thickness"])
        bcd = float(dims.get("bolt_circle_diameter", flange_od * 0.75))
        bolt_qty = int(dims.get("bolt_hole_qty", 4))
        bolt_dia = float(dims.get("bolt_hole_diameter", flange_od * 0.08))
        bore = float(dims.get("center_bore", flange_od * 0.33))
        rf_d = dims.get("raised_face_diameter")
        rf_h = dims.get("raised_face_height")

        scad = header + f"""$fn=128;

module bolt_pattern(bolt_qty, bcd, hole_d) {{
  for(i=[0:bolt_qty-1]) {{
    a = 360*i/bolt_qty;
    translate([ (bcd/2)*cos(a), (bcd/2)*sin(a), -1 ])
      cylinder(h=2000, d=hole_d);
  }}
}}

difference() {{
  cylinder(h={flange_thk}, d={flange_od});
"""
        if bore > 0:
            scad += f"  cylinder(h=2000, d={bore});\n"
        if bolt_qty > 0:
            scad += f"  bolt_pattern({bolt_qty}, {bcd}, {bolt_dia});\n"
        scad += "}\n"

        if rf_d and rf_h:
            rf_d = float(rf_d)
            rf_h = float(rf_h)
            scad += (
                f"\n// Raised face\n"
                f"translate([0,0,{flange_thk - rf_h}])\n"
                f"cylinder(h={rf_h}, d={rf_d});\n"
            )

        return scad

    # ── Block-like geometry ──
    if all(k in dims for k in ("length", "width", "height")):
        L = float(dims["length"])
        W = float(dims["width"])
        H = float(dims["height"])
        hole_qty = int(dims.get("hole_qty", 2))
        hole_d = float(dims.get("hole_diameter", 8))
        spacing = float(dims.get("hole_spacing", L * 0.5))

        scad = header + f"""$fn=96;

difference() {{
  translate([-{L/2}, -{W/2}, 0])
    cube([{L},{W},{H}]);

  for(i=[0:{hole_qty-1}]) {{
    x = -{spacing/2} + i*({spacing}/max({hole_qty-1},1));
    translate([x, 0, -1])
      cylinder(h={H+2}, d={hole_d});
  }}
}}
"""
        return scad

    # ── Fallback: simple cylinder ──
    return header + "// No recognized geometry keys found.\n"


def generate_cadquery(spec: dict) -> str:
    """Generate a CadQuery Python script from a part specification (v2 geometry engine).
    Supports flange-like (with bolt patterns, raised faces) and block-like geometries.
    """
    dims = spec.get("dimensions", {})
    name = spec.get("part_name", "part")
    part_type = spec.get("part_type", "cylinder")

    header = (
        f'"""CadQuery model for {name}\n'
        f'Item: {spec.get("item_name", "")}\n'
        f'NIIN: {spec.get("niin", "")}\n'
        f'Part type: {part_type}\n'
        f'Generated by NSN Intelligence Suite (v2 engine)\n'
        f'"""\n'
        f'import cadquery as cq\n\n'
    )

    lines = [header]

    # ── Flange-like geometry ──
    if "flange_od" in dims and "flange_thickness" in dims:
        flange_od = float(dims["flange_od"])
        flange_thk = float(dims["flange_thickness"])
        bcd = float(dims.get("bolt_circle_diameter", flange_od * 0.75))
        bolt_qty = int(dims.get("bolt_hole_qty", 4))
        bolt_dia = float(dims.get("bolt_hole_diameter", flange_od * 0.08))
        bore = float(dims.get("center_bore", flange_od * 0.33))

        lines.append(f"# Flange body")
        lines.append(f"result = cq.Workplane('XY').circle({flange_od / 2}).extrude({flange_thk})")

        if bore > 0:
            lines.append(f"\n# Center bore")
            lines.append(f"result = result.faces('>Z').workplane().hole({bore})")

        if bolt_qty > 0:
            lines.append(f"\n# Bolt hole pattern ({bolt_qty} holes on BCD {bcd}mm)")
            lines.append(f"result = (result.faces('>Z').workplane()")
            lines.append(f"          .polarArray({bcd / 2}, 0, 360, {bolt_qty})")
            lines.append(f"          .hole({bolt_dia}))")

        rf_d = dims.get("raised_face_diameter")
        rf_h = dims.get("raised_face_height")
        if rf_d and rf_h:
            rf_d = float(rf_d)
            rf_h = float(rf_h)
            lines.append(f"\n# Raised face")
            lines.append(f"result = result.faces('>Z').workplane().circle({rf_d / 2}).extrude({rf_h})")

        lines.append("")
        lines.append("# Export")
        lines.append(f'cq.exporters.export(result, "{name}.step")')
        lines.append(f'cq.exporters.export(result, "{name}.stl")')
        return "\n".join(lines)

    # ── Block-like geometry ──
    if all(k in dims for k in ("length", "width", "height")):
        L = float(dims["length"])
        W = float(dims["width"])
        H = float(dims["height"])
        hole_qty = int(dims.get("hole_qty", 2))
        hole_d = float(dims.get("hole_diameter", 8))
        spacing = float(dims.get("hole_spacing", L * 0.5))

        lines.append(f"# Block body")
        lines.append(f"result = cq.Workplane('XY').box({L}, {W}, {H}, centered=(True, True, False))")

        if hole_qty >= 1:
            lines.append(f"\n# Through holes ({hole_qty} holes, {hole_d}mm dia, {spacing}mm spacing)")
            if hole_qty <= 1:
                lines.append(f"result = result.faces('>Z').workplane().center(0, 0).hole({hole_d})")
            else:
                lines.append(f"start_x = -{spacing / 2}")
                lines.append(f"step = {spacing / max(hole_qty - 1, 1)}")
                lines.append(f"for i in range({hole_qty}):")
                lines.append(f"    x = start_x + i * step")
                lines.append(f"    result = result.faces('>Z').workplane().center(x, 0).hole({hole_d})")

        lines.append("")
        lines.append("# Export")
        lines.append(f'cq.exporters.export(result, "{name}.step")')
        lines.append(f'cq.exporters.export(result, "{name}.stl")')
        return "\n".join(lines)

    # ── Fallback ──
    lines.append("# No recognized geometry - empty workplane")
    lines.append("result = cq.Workplane('XY')")
    return "\n".join(lines)


def run_cad_generation_worker(log_id, niin):
    """Generate CAD files for a single NIIN from PUBLOG characteristics."""
    try:
        db = _worker_db()

        spec = build_part_spec_from_publog(db, niin)
        spec_json = json.dumps(spec, indent=2)

        scad_code = generate_openscad(spec)
        cq_code = generate_cadquery(spec)

        # Create output directory
        out_dir = CAD_OUTPUT_DIR / spec["part_name"]
        out_dir.mkdir(parents=True, exist_ok=True)

        # Write files
        (out_dir / "spec.json").write_text(spec_json, encoding="utf-8")
        (out_dir / "model.scad").write_text(scad_code, encoding="utf-8")
        (out_dir / "model_cadquery.py").write_text(cq_code, encoding="utf-8")

        # Write README
        readme = f"# {spec['part_name']}\n\n"
        readme += f"**Item:** {spec.get('item_name', '')}\n"
        readme += f"**NIIN:** {niin}\n\n"
        readme += f"## Files\n"
        readme += f"- `spec.json` — Part specification\n"
        readme += f"- `model.scad` — OpenSCAD source (open with OpenSCAD)\n"
        readme += f"- `model_cadquery.py` — CadQuery script (run with: `python model_cadquery.py`)\n\n"
        readme += f"## Dimensions\n"
        for k, v in spec.get("dimensions", {}).items():
            readme += f"- {k}: {v} {spec.get('units', 'mm')}\n"
        readme += f"\n## Features\n"
        for f_name in spec.get("features", []):
            readme += f"- {f_name}\n"
        readme += f"\n## Materials\n"
        for m in spec.get("materials", []):
            readme += f"- {m}\n"
        (out_dir / "README.md").write_text(readme, encoding="utf-8")

        # Record job in DB
        db.execute("""
            INSERT INTO cad_jobs (niin, status, spec_json, scad_code, cadquery_code, output_dir, finished_at)
            VALUES (?, 'completed', ?, ?, ?, ?, datetime('now'))
        """, (niin, spec_json, scad_code, cq_code, str(out_dir)))
        db.commit()
        db.close()

        _update_log(log_id, "success", 1, f"Generated CAD files for NIIN {niin} in {out_dir}")
    except Exception as e:
        log.exception("CAD generation failed for NIIN %s", niin)
        _update_log(log_id, "error", log_text=str(e))


def run_cad_batch_worker(log_id, min_chars=5, max_items=20):
    """Generate CAD files for multiple NIINs that have sufficient characteristics."""
    try:
        db = _worker_db()

        niins = db.execute("""
            SELECT pc.niin, COUNT(*) cnt
            FROM publog_characteristics pc
            INNER JOIN forecast f ON pc.niin = f.niin
            GROUP BY pc.niin
            HAVING cnt >= ?
            ORDER BY cnt DESC
            LIMIT ?
        """, (min_chars, max_items)).fetchall()

        count = 0
        for row in niins:
            niin = row["niin"]
            spec = build_part_spec_from_publog(db, niin)
            spec_json = json.dumps(spec, indent=2)
            scad_code = generate_openscad(spec)
            cq_code = generate_cadquery(spec)

            out_dir = CAD_OUTPUT_DIR / spec["part_name"]
            out_dir.mkdir(parents=True, exist_ok=True)
            (out_dir / "spec.json").write_text(spec_json, encoding="utf-8")
            (out_dir / "model.scad").write_text(scad_code, encoding="utf-8")
            (out_dir / "model_cadquery.py").write_text(cq_code, encoding="utf-8")

            db.execute("""
                INSERT OR REPLACE INTO cad_jobs (niin, status, spec_json, scad_code, cadquery_code, output_dir, finished_at)
                VALUES (?, 'completed', ?, ?, ?, ?, datetime('now'))
            """, (niin, spec_json, scad_code, cq_code, str(out_dir)))
            count += 1

        db.commit()
        db.close()
        _update_log(log_id, "success", count, f"Generated CAD models for {count} NIINs (min {min_chars} characteristics).")
    except Exception as e:
        log.exception("Batch CAD generation failed")
        _update_log(log_id, "error", log_text=str(e))


@app.route("/api/run/cad-generate", methods=["POST"])
def api_run_cad_generate():
    data = request.json if request.is_json else {}
    niin = data.get("niin")
    if niin:
        log_id = _start_pipeline("cad_generate", run_cad_generation_worker, niin)
    else:
        min_chars = int(data.get("min_chars", 5))
        max_items = int(data.get("max_items", 20))
        log_id = _start_pipeline("cad_batch", run_cad_batch_worker, min_chars, max_items)
    return jsonify({"job_id": log_id, "status": "running"})


@app.route("/api/cad/jobs")
def api_cad_jobs():
    db = get_db()
    niin = request.args.get("niin", "")
    if niin:
        rows = db.execute("SELECT * FROM cad_jobs WHERE niin=? ORDER BY created_at DESC", (niin,)).fetchall()
    else:
        rows = db.execute("SELECT id, niin, status, output_dir, created_at, finished_at FROM cad_jobs ORDER BY created_at DESC LIMIT 100").fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/cad/job/<int:job_id>")
def api_cad_job_detail(job_id):
    db = get_db()
    row = db.execute("SELECT * FROM cad_jobs WHERE id=?", (job_id,)).fetchone()
    db.close()
    if not row:
        return jsonify({"error": "not found"}), 404
    return jsonify(dict(row))


@app.route("/api/cad/preview/<niin>")
def api_cad_preview(niin):
    """Generate a preview spec without saving files."""
    db = get_db()
    spec = build_part_spec_from_publog(db, niin)
    scad = generate_openscad(spec)
    cq = generate_cadquery(spec)
    db.close()
    return jsonify({"spec": spec, "openscad": scad, "cadquery": cq})


# ──────────────────────────────────────────────
# SECTION I-D: Updated Command Center & Dashboard
# ──────────────────────────────────────────────

@app.route("/api/run/all", methods=["POST"])
def api_run_all():
    """Run decision engine (the main pipeline that uses all data)."""
    log_id = _start_pipeline("decision_engine", run_decision_engine_worker)
    return jsonify({"job_id": log_id, "status": "running"})


# ──────────────────────────────────────────────
# SECTION J: App Startup
# ──────────────────────────────────────────────

if __name__ == "__main__":
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    PDF_DIR.mkdir(parents=True, exist_ok=True)
    CAD_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    CRAWLER_DL_DIR.mkdir(parents=True, exist_ok=True)
    init_db()
    seed_default_config()
    log.info("NSN Intelligence Suite starting on http://127.0.0.1:5000")
    app.run(host="127.0.0.1", port=5000, debug=True)
